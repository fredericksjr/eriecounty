#!/usr/bin/env python

"""convertNASReports.py: read in heap autopay txt file file and export as single record text file (outAutopay.txt) """

__author__   = "Frederick Sulkowski"
__email__   = "frederick.sulkowski@erie.gov"


from argparse import ArgumentDefaultsHelpFormatter
from ftplib import ftpcp
from turtle import goto
import os
import pandas as pd
pd.io.formats.excel.ExcelFormatter.header_style= None
from tkinter import messagebox

import smtplib #fms
from email.mime.text import MIMEText #fms
from email.mime.multipart import MIMEMultipart #fms
from email.mime.application import MIMEApplication #fms
# Define sender and receiver email addresses
sender = 'frederick.sulkowski@erie.gov'
recipient = 'fsulkowski@hotmail.com'


def main():

    # read files in current directory
    for f_name in os.listdir('.'):
        if (f_name.startswith('001-X-WMMRELNA-14') and f_name.endswith('.txt')) or (f_name.startswith('001-X-WMMRRJNA-14') and f_name.endswith('.txt')) :
            wmmrelna(f_name)

        if f_name.startswith('001-X-WMMREXNA-14') and f_name.endswith('.txt') :
            wmmrexna(f_name)
        if f_name.startswith('001-X-WM4DELNA-14') and f_name.endswith('.txt') :
            wm4delna(f_name)
        if f_name.startswith('001-X-WM4DEXNA-14') and f_name.endswith('.txt') :
            wm4dexna(f_name)
        if f_name.startswith('001-X-14-WMSCSDX5') and f_name.endswith('.txt') :
            wmscsdx5(f_name)
        if (f_name.startswith('001-X-14-WMSBF630') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSBP630') and f_name.endswith('.txt')) :
            wmsbf630(f_name)
        if (f_name.startswith('001-X-14-WMSBF650') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSBP650') and f_name.endswith('.txt')) :
            wmsbf650(f_name)
        if (f_name.startswith('001-X-14-WMSBF650') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSBP650') and f_name.endswith('.txt')) :
            wmsbf650(f_name)
        if (f_name.startswith('001-X-14-WMSA5257') and f_name.endswith('.txt')) :
            wmsa5257(f_name)
        if (f_name.startswith('001-X-14-WMSC1025') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1040') and f_name.endswith('.txt')) :
            wmsc1025(f_name)
        if (f_name.startswith('001-X-14-WMSC1026') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1041') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1661') and f_name.endswith('.txt')) :
            wmsc1026(f_name)
        if (f_name.startswith('001-X-14-WMSC2027') and f_name.endswith('.txt')) :
            wmsc2027(f_name)            
        if (f_name.startswith('001-X-14-WMSBHIGH') and f_name.endswith('.txt')) :
            wmsbhigh(f_name)            
        if (f_name.startswith('001-X-14-WMSC4210') and f_name.endswith('.txt')) :
            wmsc4210(f_name)            
        if (f_name.startswith('WINR0799_PDF') and f_name.endswith('.txt')) :
            winr0799(f_name)            
        if (f_name.startswith('CSPROD-BBP101PRTF') and f_name.endswith('.txt')) :
            wbb101(f_name)            
        if (f_name.startswith('PERIE-PYCHECKS') and f_name.endswith('.txt')) :
            lfrr0011(f_name)            
        if (f_name.startswith('PERIE-SVRETVCH') and f_name.endswith('.txt')) :            
            retrt010(f_name)
        if (f_name.startswith('001-X-14-WMSC1047') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1662') and f_name.endswith('.txt')):
            wmsc1047(f_name)
        if (f_name.startswith('PERIE-DPESROLL0000') and f_name.endswith('.txt')) :
            pabs4002(f_name)
        if (f_name.startswith('PERIECNB560-CNS') and f_name.endswith('.txt')) :
            cns00125(f_name)

            

#===================================================================================================

def wmmrelna(f_name):

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "OLD DEFCT\t" + "NEW DEFCT\t" + "NET CHNG\t" +
                "SPEC ALERT\t" + "OLD ETLMT\t" + "NEW ETLMT\t" + "NET CHNG\t" + "REJ/EXCEPT")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            if 'WMMRRJNA' in f_name:
                office = record[54:59].strip() + '\t'
                unit = record[65:70].strip() + '\t'
                worker = record[80:86].strip() + '\t'
            else:
                office = record[44:49].strip() + '\t'
                unit = record[56:61].strip() + '\t'
                worker = record[72:78].strip() + '\t'

        if (not line.startswith('WMRIQ5') and not line.startswith('DISTRICT') and not line.startswith('OLD DEFCT') 
            and not line.startswith('CASE NAME') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[26:47].strip() + '\t'
                caseNumber = record[48:58].strip() + '\t'
                oldDefct = record[60:67].strip() + '\t'
                newDefct = record[72:78].strip() + '\t'
                netChng1 = record[82:89].strip() + '\t'
                specAlert = record[94:].strip() + '\t'

            if countLines == 2:
                oldEtlmt = record[60:67].strip() + '\t'
                newEtlmt = record[71:78].strip() + '\t'
                netChng2 = record[82:89].strip() + '\t'
                rejExcept = record[94:]
                endPos = rejExcept.find(carriage)
                rejExcept = rejExcept[0:endPos].strip()
                out_str = (office + unit + worker + caseName + caseNumber + oldDefct + newDefct + netChng1 + specAlert + 
                           oldEtlmt + newEtlmt + netChng2 + rejExcept)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    if 'WMMRRJNA' in f_name:
        messagebox.showinfo("Completed - WMMRRJNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + "!")
    else:
        messagebox.showinfo("Completed - WMMRELNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + "!")

    return

#===================================================================================================

def wmmrexna(f_name):

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "EXCEPTED REASON")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            office = record[54:59].strip() + '\t'
            unit = record[65:70].strip() + '\t'
            worker = record[80:86].strip() + '\t'

        if (not line.startswith('WMRIQ4') and not line.startswith('DISTRICT') and not line.startswith('CASE NAME') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[28:55].strip() + '\t'
                caseNumber = record[56:66].strip() + '\t'
                exceptedReason = record[84:].strip()
                out_str = (office + unit + worker + caseName + caseNumber + exceptedReason)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMMREXNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wm4delna(f_name):

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "MESSAGE #1\t" + "PREV 2\t" +
                "PREV 1\t" + "CUR IVD\t" + "NEW FS\t" + "CUR FA\t" + "SPEC ALERTS\t" + "MESSAGE #2\t" + "CUM 2\t" +
                "CUM 1\t" + "CUR CUM\t" + "OLD FS\t" +"PRE FA\t" + "ACTION CODE\t" + "MONTH OBL\t" + "EXMT 2\t" +
                "EXMT 1\t" + "EXMT CUR")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            office = record[47:52].strip() + '\t'
            unit = record[58:64].strip() + '\t'
            worker = record[75:81].strip() + '\t'

        if (not line.startswith('WMRIV2') and not line.startswith('DISTRICT') and not line.startswith('CASE NAME')
            and not line.startswith('CASE NUMBER') and not line.startswith('MONTH OBL') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[9:39].strip() + '\t'
                message1 = record[39:50].strip() + '\t'
                prev2 = record[54:61].strip() + '\t'
                prev1 = record[65:72].strip() + '\t'
                curIvd = record[76:83].strip() + '\t'
                newFs = record[87:94].strip() + '\t'
                curFa = record[98:105].strip() + '\t'
                specAlerts = record[109:].strip() + '\t'
                
            if countLines == 2  :
                caseNumber = record[14:24].strip() + '\t'
                message2 = record[39:50].strip() + '\t'
                cum2 = record[54:61].strip() + '\t'
                cum1 = record[65:72].strip() + '\t'
                curCum = record[76:83].strip() + '\t'
                oldFs = record[87:94].strip() + '\t'
                preFa = record[98:105].strip() + '\t'
                actionCode = record[109:].strip() + '\t'

            if countLines == 3  :
                monthObl = record[39:50].strip() + '\t'
                exmt2 = record[54:61].strip() + '\t'
                exmt1 = record[65:72].strip() + '\t'
                exmtCur = record[76:83].strip()
                out_str = (office + unit + worker + caseName + caseNumber + message1 + prev2 + prev1 + curIvd + newFs +
                           curFa + specAlerts + message2 + cum2 + cum1 + curCum + oldFs + preFa + actionCode + monthObl +
                           exmt2 + exmt1 + exmtCur)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WM4DELNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wm4dexna(f_name):
 
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "WMRIV1 DEXNA MRBA - EXCEPTION LIST FOR IVD [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "MESSAGE #1\t" + "PREV 2\t" +
                "PREV 1\t" + "CUR IVD\t" + "CUR FA\t" + "EXCEPTION REASON\t" + "MESSAGE #2\t" + "CUM 2\t" + "CUM 1\t" +
                "CUR CUM\t" + "PRE FA\t" +"ACTION REQUIRED\t" + "MONTH OBL\t" + "EXMT 2\t" + "EXMT 1\t" + "EXMT CUR")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            office = record[47:52].strip() + '\t'
            unit = record[58:64].strip() + '\t'
            worker = record[75:81].strip() + '\t'

        if (not line.startswith('WMRIV1') and not line.startswith('DISTRICT') and not line.startswith('CASE NAME')
            and not line.startswith('CASE NUMBER') and not line.startswith('MONTH OBL') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[9:39].strip() + '\t'
                message1 = record[39:50].strip() + '\t'
                prev2 = record[54:61].strip() + '\t'
                prev1 = record[65:72].strip() + '\t'
                curIvd = record[76:83].strip() + '\t'
                curFa = record[87:93].strip() + '\t'
                exceptionReason = record[98:].replace(nullChar,' ').strip() + '\t'

            if countLines == 2  :
                caseNumber = record[14:24].strip() + '\t'
                message2 = record[39:50].strip() + '\t'
                cum2 = record[54:61].strip() + '\t'
                cum1 = record[65:72].strip() + '\t'
                curCum = record[76:83].strip() + '\t'
                preFa = record[87:93].strip() + '\t'
                actionRequired = record[98:].strip() + '\t'

            if countLines == 3  :
                monthObl = record[39:50].strip() + '\t'
                exmt2 = record[54:61].strip() + '\t'
                exmt1 = record[65:72].strip() + '\t'
                exmtCur = record[76:83].strip()
                out_str = (office + unit + worker + caseName + caseNumber + message1 + prev2 + prev1 + curIvd +
                           curFa + exceptionReason + message2 + cum2 + cum1 + curCum + preFa + actionRequired +
                           monthObl + exmt2 + exmt1 + exmtCur)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
               
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    messagebox.showinfo("Completed - WM4DEXNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmscsdx5(f_name):
    
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0
    statusNext = ' '

    outfile = "ec_" + f_name
    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfile):
        os.remove(outfile)
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("STATUS\t" + "CASE TYPE\t" + "SSN\t" + "LAST NAME\t" + "FIRST NAME\t" + "M I\t" + "CASE NUMBER\t" +
                "CIN\t" + "OLD ETMNT\t" + "NEW ETMNT\t" + "NEW ALMNT\t" + "F T\t" + "AUTH FROM\t" + "AUTH TO\t" + "TRANSACTION DISPOSITION\t" +
                "TRANSACTION DISPOSITION\t" + "TRANSACTION DISPOSITION\t" + "TRANSACTION DISPOSITION")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if statusNext == 'Y' and not line.startswith(escape):
            status = record[0:64].strip() + '\t'
            statusNext = ' '

        if line.startswith('TYPE'):
            statusNext = 'Y'
        
        if (not line.startswith('DISTRICT') and not line.startswith('AUTO SDX') and not line.startswith('AUTOMATED')
            and not line.startswith('TRANSACTION CONTROL REPORT') and not line.startswith('CASE') 
            and not line.startswith('TYPE') and not line.startswith('EXCEPTION') and not line.startswith('ELIGIBLE')
            and not line.startswith('RFC') and not line.startswith('NON WORKER') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0) :

            if countLines == 3 :
                countLines = 0

            if countLines == 2 :
                recordCheck = record[0:20].strip()

                if recordCheck == "" :
                    transactionDisposition3 = record[26:60].strip() + '\t'
                    transactionDisposition4 = record[61:95].strip()
                else :
                    transactionDisposition3 = ' ' + '\t'
                    transactionDisposition4 = ' '
                    countLines = 0

                out_str = (status + caseType + ssn + lastName + firstName + mi + caseNumber + cin + oldEtmnt +
                    newEtmnt + newAlmnt + ft + authFrom + authTo + transactionDisposition1 + transactionDisposition2 +
                    transactionDisposition3 + transactionDisposition4) 
                           
            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 

            countLines = countLines + 1

            if countLines == 1 :
                caseType = record[0:4].strip() + '\t'
                if record[7:8].strip() == 'd' :
                    record2 = '       d           ' + record[11:]
                    record = record2
                ssn = record[7:16].strip() + '\t'
                lastName = record[19:39].strip() + '\t'
                firstName = record[39:50].strip() + '\t'
                mi = record[50:51].strip() + '\t'
                caseNumber = record[54:64].strip() + '\t'
                cin = record[67:75].replace(nullChar,' ').strip() + '\t'
                oldEtmnt = record[78:86].strip() + '\t' 
                newEtmnt = record[89:97].strip() + '\t'
                newAlmnt = record[100:107].strip() + '\t'
                ft = record[110:111].strip() + '\t'
                authFrom = record[114:122].strip() + '\t'
                authTo  = record[122:130].strip() + '\t'

            if countLines == 2 :
                transactionDisposition1 = record[26:60].strip() + '\t'
                transactionDisposition2 = record[61:95].strip() + '\t'

    out_str = (status + caseType + ssn + lastName + firstName + mi + caseNumber + cin + oldEtmnt +
        newEtmnt + newAlmnt + ft + authFrom + authTo + transactionDisposition1 + transactionDisposition2 +
        transactionDisposition3 + transactionDisposition4)
                           
    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
        countWrites = countWrites +1 

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WM4DEXNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsbf630(f_name):
    
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0
    heapPerson = ' '

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#fms    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"

    if f_name.startswith('001-X-14-WMSBF630'):    
        outfilexlsx = "WINR0630 HEAP-SNAP Mass Auth Exception [" + os.path.splitext(f_name[28:39])[0] + "].xlsx"
    if f_name.startswith('001-X-14-WMSBP630'):
        outfilexlsx = "WINR0630 HEAP-TA Mass Auth Exception [" + os.path.splitext(f_name[28:39])[0] + "].xlsx"
        
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)
  
    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("CASE NUMBER\t" + "CASE NAME\t" + "LASTNAME\t" + "FIRSTNAME\t" + "MI\t" + "EXCEPTION REASON")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('DISTRICT') and not line.startswith('LIST OF')
            and not line.startswith('CASE NUMBER') and not line.startswith(escape)) :
            # and not line.startswith(carriage) and len(line) != 0) :

                    countLines = countLines + 1
                    lengthOfLine = len(line)

            # format 3 lines of data into one            
                    if line[0] != '\"':
                        line = ("\"" + line + "\"")

                    heapPerson = heapPerson + line

                    if countLines == 1 and lengthOfLine < 118:
                        heapPerson = heapPerson[:116] + "   \""

                    if countLines == 1:
                        new_str = heapPerson.replace("\"", " ")

            # case number (caseNumber)
                        caseNumber = new_str[2:13].strip() + '\t'
            # case name (caseName)                
                        if "." in new_str[15:51]:
                            new_str = new_str[:15] + new_str[15:51].replace("."," ") + new_str[51:]

                        if ", " in new_str[15:51] or " ," in new_str[15:51]:
                            new_str = new_str.replace(", ",",")
                            new_str = new_str.replace(" ,",",")
                    
                        tempName = new_str[15:50].strip()
                        caseName = new_str[15:50].strip() + '\t'


#                        def get_name_parts(name):
                        comma_split = tempName.split(',')
                        tempLast = comma_split[0]
                        caseLast = str(tempLast) + '\t'

                        try:
                            tempFirst = comma_split[1].split(',')[0]
                            caseFirst = str(tempFirst) + '\t'
                        except IndexError:
                            caseFirst = ' \t'

                        try:
                            tempMI = comma_split[2].split(',')[0]
                            caseMI = str(tempMI) + '\t'
                        except IndexError:
                            caseMI = ' \t'

            # exception reason (exceptionReason)
                        exceptionReason = new_str[52:76].strip()


               
                        out_str = caseNumber + caseName + caseLast + caseFirst + caseMI + exceptionReason

            # write to results file                                
                        if out_str != ' ':
                            f = open(outfile,"a")
                            f.write(out_str + "\n")
                            f.close()

                            countWrites = countWrites +1 
                            countLines = 0
                            heapPerson = ' '

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None,low_memory=False)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSBF630", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsbf650(f_name):
    
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"

    if f_name.startswith('001-X-14-WMSBF650'):    
        outfilexlsx = "WINR0650 HEAP-SNAP Mass Auth Eligible [" + os.path.splitext(f_name[28:39])[0] + "].xlsx"
    if f_name.startswith('001-X-14-WMSBP650'):
        outfilexlsx = "WINR0650 HEAP-TA Mass Auth Eligible [" + os.path.splitext(f_name[28:39])[0] + "].xlsx"

    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    # read file
    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("CASE NUMBER\t" + "CASE NAME\t" + "ADDRESS\t" + "CITY\t" + "STATE\t" + "ZIP\t" + "Local Office\t" + "UNIT\t" + 
                            "WORKER\t" + "ACT (Local Action Code)\t" + "PAY TYPE\t" + "Method\t" + "AMOUNT\t" + "IS (Issuance Code)\t" + 
                            "SC (Schedule Code)\t" + "PU (Pick Up Code)\t" + "PERIOD FROM\t" + "PERIOD TO\t" + "FUEL TYPE\t" +
                            "SPC CLM (Special Claim\t)" + "LOCAL USAGE\t" + "VP (Voucher Payment\t" + "HH SIZE\t" + "INCOME\t" +
                            "HEAP VENDOR ID\t" + "CUSTOMER ACCT NO\t" + "NX (Heap Nominal Benefit\t" + "CASE TYPE")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

        # setup variables
        heapPerson = ' '
    
        for line in lines:
            line = line.strip().upper()
            line = line.strip()

            if (not line.startswith('CASE NAME') and not line.startswith('DISTRICT') and not line.startswith('ADDRESS') 
                and not line.startswith('CASE NUMBER') and not line.startswith('RC132') and not line.startswith('* * *  DIS') 
                and not line.startswith('"*  TOTAL') and not line.startswith('*  TOTAL') and not line.startswith("\u001B")):

                countLines = countLines + 1
                lengthOfLine = len(line)

                # format 3 lines of data into one            
                if line[0] != '\"':
                    line = ("\"" + line + "\"")

                heapPerson = heapPerson + line

                if countLines == 1 and lengthOfLine < 118:
                    heapPerson = heapPerson[:116] + "   \""

    #            if countLines == 2 and heapPerson[158] == ' ':
    #                heapPerson = heapPerson[:158] + heapPerson[160:]

                if countLines == 3 and heapPerson[188] == '"':
                    heapPerson = heapPerson[:189] + '           ' + heapPerson[189:]

                if countLines == 3:
                    new_str = heapPerson.replace("\"", " ")
                    
                    caseNumber = new_str[2:14].strip() + '\t'
                    caseName = new_str[15:51].strip() + '\t'
                    caseAddress = new_str[121:157].strip() + '\t'
                    caseCity = new_str[201:217].strip() + '\t'
                    caseState = new_str[219:221].strip() + '\t'
                    caseZip = new_str[225:230].strip() + '\t'
                    localOffice = new_str[52:55].strip() + '\t'
                    caseUnit = new_str[158:163].strip() + '\t'
                    caseWorker = new_str[238:243].strip() + '\t'
                    localActionCode = new_str[60:61].strip() + '\t'
                    payType = new_str[64:66].strip() + '\t'
                    caseMethod = new_str[70:72].strip() + '\t'
                    caseAmount = new_str[73:82].strip() + '\t'
                    issuanceCode = new_str[84:85].strip() + '\t'
                    scheduleCode = new_str[87:87].strip() + '\t'
                    pickUpCode = new_str[90:91].strip() + '\t'
                    periodFrom = new_str[94:100].strip() + '\t'
                    periodTo = new_str[101:107].strip() + '\t'
                    fuelType = new_str[109:110].strip() + '\t'
                    spcClm = new_str[114:115].strip() + '\t'
                    localUsage = new_str[118:118].strip() + '\t'
                    voucherPayment = new_str[179:180].strip() + '\t'
                    hhSize = new_str[186:188].strip() + '\t'
                    caseIncome = new_str[189:199].strip() + '\t'
                    heapVendorId = new_str[250:257].strip() + '\t'
                    customerAcctNo = new_str[270:285].strip() + '\t'
                    heapNominalBenefit = ' ' + '\t'
                    caseType = new_str[309:311].strip()

                    out_str = (caseNumber + caseName + caseAddress + caseCity + caseState + caseZip + localOffice + caseUnit + caseWorker + 
                        localActionCode + payType + caseMethod + caseAmount + issuanceCode + scheduleCode + pickUpCode + periodFrom + periodTo + fuelType +
                        spcClm + localUsage + voucherPayment + hhSize + caseIncome + heapVendorId + customerAcctNo + heapNominalBenefit + caseType)

            # write to results file                                
                    if out_str != ' ':
                        f = open(outfile,"a")
                        f.write(out_str + "\n")
                        f.close()

                        countWrites = countWrites +1 
                        countLines = 0
                        heapPerson = ' '



    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None,low_memory=False)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSBF630", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsa5257(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "INDIVIDUAL NAME\t" + "CIN\t" +
        "RETURN DATE")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()
               
        if line.startswith('LOCAL OFFICE'):
            localOffice = line[15:18].strip() + '\t'
            unit = line[37:42].strip() + '\t'
            worker = line[61:66].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*') and not line.startswith('QI-1') 
            and not line.startswith('PERIOD COVERED BY') and not line.startswith('REFERENCE') and not line.startswith('DISTRICT')
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[21:57].strip() + '\t'
                individualName = line[59:97].strip() + '\t'
                cin = line[99:107].strip() + '\t'
                returnDate = line[117:125].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + individualName + cin + returnDate)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSA5257", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc1025(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "PAY AMOUNT\t" + "PAY PERIOD")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[15:18].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('MASS AUTHORIZATION') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[23:55].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                caseType = line[56:58].strip() + '\t'
                payAmount = line[67:75].strip() + '\t'
                payPeriod = line[82:101].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + payee + caseType + payAmount + payPeriod)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC1025/1040", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc1026(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    if f_name.startswith('001-X-14-WMSC1026')  :
        outfilexlsx = "WINR1026 - MASS AUTH OF SUPPLEMENTAL SNAP - EXCEPTION [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
        out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "EXCEPTION REASON")
    if f_name.startswith('001-X-14-WMSC1041') :
        outfilexlsx = "WINR1041 - MASS AUTH OF SNAP - EXCEPTION REPORT [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
        out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "EXCEPTION REASON")
    if f_name.startswith('001-X-14-WMSC1661') :
        outfilexlsx = "WINR1661 - AUTH OF HEAP J9 SUPPL. BENEFITS - EXCEPTION [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
        out_str = ("Local Office\t" + "Unit\t" + "Worker\t" + "Case Number\t" + "Case Name\t" + "Case Type\t" +
        "Exception Reason")

    #outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

#    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
#        "EXCEPTION REASON")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[14:18].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('MASS AUTHORIZATION') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[13:45].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                caseType = line[46:48].strip() + '\t'
                exceptionReason = line[68:].strip()

                if f_name.startswith('001-X-14-WMSC1661') :
                    out_str = (localOffice + unit + worker + caseNumber + caseName + caseType + exceptionReason)
                else :
                    out_str = (localOffice + unit + worker + caseNumber + caseName + payee + caseType + exceptionReason)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC1026/1041", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc2027(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WINR2027 MASS ISSUANCE OF P-EBT - ELIGIBLE REPORT [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "PAY AMOUNT\t" + "PAY PERIOD FROM\t" + "PAY PERIOD TO\t" + "AVAIL DT")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[15:18].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('MASS ISSUANCE') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[23:55].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                caseType = line[56:58].strip() + '\t'
                payAmount = line[67:75].strip() + '\t'
                payPeriodFrom = line[82:90].strip() + '\t'
                payPeriodTo = line[93:101].strip() + '\t'
                availDt = line[107:117].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + payee + caseType + payAmount + payPeriodFrom
                    + payPeriodTo + availDt)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC2027", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsbhigh(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WMSBHIGH - RFI HIGH RISK LISTING [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE TYPE\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "INCOME\t" +
        "RESOURCE")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('DISTRICT') and not line.startswith('OFFICE') 
            and not line.startswith('RFI HIGH') and not line.startswith('UNRESOLVED')
            and not line.startswith('SORT SEQUENCE') and not line.startswith('PLEASE') 
            and not line.startswith('THE ATTACHED') and not line.startswith('THE UNRESOLVED')
            and not line.startswith('FOR QUESTIONS') and not line.startswith('ASSISTANCE')
            and not line.startswith('HITS  THAT') and not line.startswith('REVENUE')
            and not line.startswith('BUSINESS') and not line.startswith('45 DAYS')
            and not line.startswith('"HITS"') and not line.startswith('OF THE WRM.')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            if countLines == 1  :
                localOffice = record[5:8].strip() + '\t'
                unit = record[14:19].strip() + '\t'
                worker = record[24:29].strip() + '\t'
                caseType = record[36:38].strip() + '\t'
                caseNumber = record[52:62].strip() + '\t'
                caseName = record[66:96].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                income = record[98:105].strip() + '\t'
                resource = record[110:120].strip()

                out_str = (localOffice + unit + worker + caseType + caseNumber + caseName + payee + income + resource)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC2027", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc4210(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WINR4210 HEAP AUTO-CLOSE EXCEPTION REPORT [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "EXCEPTION REASON")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('REPORT DATE') and not line.startswith('*')
            and not line.startswith('DISTRICT') and not line.startswith('CASE NAME') 
            and not line.startswith('TOTAL CASES') and not line.startswith('END OF REPORT')
            and not line.startswith('REFERENCE NO')  
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            if countLines == 1  :
                localOffice = record[14:17].strip() + '\t'
                unit = record[38:43].strip() + '\t'
                worker = record[58:63].strip() + '\t'

            if countLines == 2  :
                caseName = record[0:32].strip() + '\t'
                caseNumber = record[32:42].strip() + '\t'
                exceptionReason = record[50:].strip()
#                if '/PAYEE' in caseName :
#                    caseName = caseName.replace('/PAYEE','')
#                    payee = 'TRUE\t'
#                else :
#                    payee = '\t'

                out_str = (localOffice + unit + worker + caseName + caseNumber + exceptionReason)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WINR4210", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def winr0799(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WINR0799 Upstate SSN Auto Close Monthly Exception [" + os.path.splitext(f_name[20:31])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "CASE TYPE\t" + "CID ID\t" + "VR IND\t" +
        "Date of Death\t" + "Exception Message (Case is not closed)")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        #line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('NEW YORK STATE') and not line.startswith('Upstate Welfare')
            and not line.startswith('UPSTATE AUTO') and not line.startswith('Exception List') 
            and not line.startswith('REPORT DATE') and not line.startswith('WRTS REPORT')
            and not line.startswith('THIS REPORT') and not line.startswith('CONFIDENTIAL')
            and not line.startswith('FOR RESTRICTED') and not line.startswith('USE ONLY')
            and not line.startswith('District') and not line.startswith('Case Number')
            and not line.startswith('(Case is not') and not line.startswith('Total') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            #if countLines == 1  :
            if line.startswith('Unit')  :
                localOffice = record[22:25].strip() + '\t'
                unit = record[26:31].strip() + '\t'
                worker = record[32:37].strip() + '\t'

            #if countLines == 2  :
            else  :
                if record[8:57].strip() == ''  :
                    vrInd = record[82:84].strip() + '\t'
                    exceptionMessage = record[104:].strip()
                else  :
                    caseNumber = record[8:18].strip() + '\t'
                    caseName = record[27:57].strip() + '\t'
                    caseType = record[57:59].strip() + '\t'
                    cinId = record[67:75].strip() + '\t'
                    vrInd = record[82:84].strip() + '\t'
                    dateOfDeathMM = record[90:92].strip() + '/'
                    dateOfDeathDD = record[92:94].strip() + '/'
                    dateOfDeathYYYY = record[94:98].strip()
                    if dateOfDeathYYYY == ''  :
                        dateOfDeath = '\t'
                    else  :
                        dateOfDeath = dateOfDeathMM + dateOfDeathDD + dateOfDeathYYYY + '\t'
                    exceptionMessage = record[104:].strip()

                out_str = (unit + localOffice + worker + caseNumber + caseName + caseType + cinId + 
                    vrInd + dateOfDeath + exceptionMessage)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WINR0799", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wbb101(f_name):

    escape = '\u001B'
    carriage = '\r'
    colon = '\ua789'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "FBI FS AND PA" + colon + " FS EXTRACT" + colon + " FEDERAL BENEFIT INCREASE [" + os.path.splitext(f_name[30:40])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "Mailing Address Line 1\t" +
        "Mailing Address City\t" + "Mailing Address State\t" + "Mailing Address Zip\t" + "SI Benefits\t" + "Veterans Benefits\t" +
        "Railroad Retirement Benefits\t" + "SS Benefits\t" + "EST EXEC")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        #line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('1234567890') and not line.startswith('File Sequence')
            and not line.startswith('WW')
            and not line.startswith('Product') and not line.startswith('Printer') 
            and not line.startswith('-Host') and not line.startswith('ALBEU')
            and not line.startswith('Host Full') and not line.startswith('Data File')
            and not line.startswith('Characters') and not line.startswith('Lines Per')
            and not line.startswith('Top Margin') and not line.startswith('Bottom Margin')
            and not line.startswith('Warning:') and not line.startswith('legally') 
            and not line.startswith('groups') and not line.startswith('government')
            and not line.startswith('constitutes') and not line.startswith('notice') 
            and not line.startswith('REPORT') and not line.startswith('WELFARE')
            and not line.startswith('FS EXTRACT') and not line.startswith('DISTRICT')
            and not line.startswith('CASE NAME')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            #if countLines == 1  :
            if line.startswith('OFFICE')  :
                localOffice = record[13:18].strip() + '\t'
                unit = record[24:29].strip() + '\t'
                worker = record[39:44].strip() + '\t'

            #if countLines == 2  :
            else  :
                if record[6:36].strip() != ''  :
                    if record[5:7].strip() == '' :
                        caseName = record[7:37].strip() + '\t'
                        caseNumber = record[41:53].strip() + '\t'
                        mailingAddressLine1 = record[56:96].strip() + '\t'
                        siBenefits = record[98:100].strip() + '\t'
                        veteranBenefits = record[101:103].strip() + '\t'
                        railroadRetirementBenefits = record[104:106].strip() + '\t'
                        ssBenefits = record[107:110].strip() + '\t'
                        estExec = record[112:].strip()
                    else :    
                        caseName = record[6:36].strip() + '\t'
                        caseNumber = record[40:52].strip() + '\t'
                        mailingAddressLine1 = record[55:95].strip() + '\t'
                        siBenefits = record[97:99].strip() + '\t'
                        veteranBenefits = record[100:102].strip() + '\t'
                        railroadRetirementBenefits = record[103:105].strip() + '\t'
                        ssBenefits = record[106:109].strip() + '\t'
                        estExec = record[111:].strip()
                else  :
                    mailingAddressCity = record[56:72].strip() + '\t'
                    mailingAddressState = record[73:75].strip() + '\t'
                    mailingAddressZip = record[77:82].strip() + '\t'

                    out_str = (localOffice + unit + worker + caseName + caseNumber + mailingAddressLine1 + mailingAddressCity + 
                        mailingAddressState + mailingAddressZip + siBenefits + veteranBenefits + railroadRetirementBenefits +
                        ssBenefits + estExec)

    # write to results file                                
                    if out_str != ' ':
                        f = open(outfile,"a")
                        f.write(out_str + "\n")
                        f.close()
                        countWrites = countWrites +1 
                        countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WBB101", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def lfrr0011(f_name):

    import openpyxl as pxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    #import pandas as pd
    #import pandas.io.formats.style
#    import Jinja2
    from decimal import Decimal

#    pandas.io.formats.excel.ExcelFormatter.header_style= None

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True
    startSummary = False
    getFileName = False

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)
    outfile2 = "ec2_" + f_name
    if os.path.exists(outfile2):
        os.remove(outfile2)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("CHECK/SERV PERIOD Start\t" + "CHECK/SERV PERIOD End\t" + "CHECK NUMBER\t" + "Check Date\t" + "Payment Amount\t" + "Prior Year Amount\t" +
        "Current Year Amount\t" + "Category\t" + "Check Status\t" + "Cancel Date\t" + "Check Type\t" + "Case #\t" + "Adjustment Memo")
    out_str2 = ("CATEGORY\t" + "Prior Year ISSUE AMOUNT\t" + "Prior Year CANCEL AMOUNT\t" + "Prior Year BALANCE\t" + 
        "CURRENT YEAR ISSUE AMOUNT\t" + "CURRENT YEAR CANCEL AMOUNT\t" + "CURRENT YEAR BALANCE")    

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    if out_str2 != ' ':
        f = open(outfile2,"a")
        f.write(out_str2 + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('REPORT PERIOD') and getFileName == False :
            outfilexlsx = "LFRR0011 PAYMENTS FOR PRIOR YEAR SERVICES [" + record[15:34].replace('/','') + "].xlsx"
            getFileName = True
            if os.path.exists(outfilexlsx):
                os.remove(outfilexlsx)

        if line.startswith('CATEGORY') :
            startSummary = True

        if (not line.startswith('REPORT') and not line.startswith('BICS') and not line.startswith('SORTED')
            and not line.startswith('PAYMENTS') and not line.startswith('CHECK') and not line.startswith('*****')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startSummary == False):

            countLines = countLines + 1

#            if countLines == 1  :
            checkServPeriodStart = record[0:8].strip() + '\t'
            checkServPeriodEnd = record[11:19].strip() + '\t'
            checkNumber = record[21:33].strip() + '\t'
            checkDate = record[35:43].strip() + '\t'
            paymentAmount = record[44:52].strip() + '\t'
            priorYearAmount = record[53:61].strip() + '\t'
            currentYearAmount = record[62:70].strip() + '\t'
            category = record[72:81].strip() + '\t'
            checkStatus = record[82:85].strip() + '\t'
            cancelDate = record[87:95].strip() + '\t'
            checkType = record[97:100].strip() + '\t'
            caseNumber = record[102:113].strip() + '\t'
            adjustmentMemo = record[114:].strip()



            out_str = (checkServPeriodStart + checkServPeriodEnd + checkNumber + checkDate + paymentAmount + priorYearAmount +
                currentYearAmount + category + checkStatus + cancelDate + checkType + caseNumber + adjustmentMemo)

            # write to results file                                
            if out_str != ' ':
                f = open(outfile,"a")
                f.write(out_str + "\n")
                f.close()
                countWrites = countWrites +1 
                countLines = 0

        if (not line.startswith('TOTALS') and not line.startswith('CATEGORY')
            and not line.startswith(escape) and not line.startswith(carriage) 
            and len(line) != 0 and startSummary == True):

            countLines = countLines + 1

#            if countLines == 1  :
            sumCategory = record[0:8].strip() + '\t'
            sumPriorYearIssueAmount = record[12:24].strip() + '\t'
            sumPriorYearCancelAmount = record[28:40].strip() + '\t'
            sumPriorYearBalance = record[46:58].strip() + '\t'
            sumCurrentYearIssueAmount = record[83:95].strip() + '\t'
            sumCurrentYearCancelAmount = record[99:111].strip() + '\t'
            sumCurrentYearBalance = record[116:128].strip()

            out_str2 = (sumCategory + sumPriorYearIssueAmount + sumPriorYearCancelAmount + sumPriorYearBalance + 
                sumCurrentYearIssueAmount + sumCurrentYearCancelAmount + sumCurrentYearBalance)

            # write to results file                                
            if out_str2 != ' ':
                f = open(outfile2,"a")
                f.write(out_str2 + "\n")
                f.close()
                countWrites = countWrites +1 
                countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=0)
    df['CHECK/SERV PERIOD Start'] = pd.to_datetime(df['CHECK/SERV PERIOD Start']).dt.strftime('%m/%d/%Y')
    df['CHECK/SERV PERIOD End'] = pd.to_datetime(df['CHECK/SERV PERIOD End']).dt.strftime('%m/%d/%Y')
    df['Check Date'] = pd.to_datetime(df['Check Date']).dt.strftime('%m/%d/%Y')
    #df['Payment Amount'] = df['Payment Amount'].map('{:,.2f}'.format)
    #df['Payment Amount'] = df['Payment Amount'].map('{:,.2f}'.format)
    #values = pd.to_numeric(df['Payment Amount].'].str.replace("[.%]", "").str.replace(",", "."))
    #df= df['Payment Amount'].str.replace(',', '').astype(float)
    df['Payment Amount'] = pd.to_numeric(df['Payment Amount'])
#    df['Payment Amount'] = df['Payment Amount'].apply(lambda x: format(float(x),".2f"))
#    df['Payment Amount'] = df['Payment Amount'].
    #df['Prior Year Amount'] = df['Prior Year Amount'].map('{:,.2f}'.format)
    df['Prior Year Amount'] = pd.to_numeric(df['Prior Year Amount'])
    #df['Current Year Amount'] = df['Current Year Amount'].map('{:,.2f}'.format)
    df['Current Year Amount'] = pd.to_numeric(df['Current Year Amount'])
    df['Cancel Date'] = pd.to_datetime(df['Cancel Date']).dt.strftime('%m/%d/%Y')
    df.to_excel(outfilexlsx,'Detail',index=False,header=True)           
    os.remove(outfile)            

    excel_book = pxl.load_workbook(outfilexlsx)

    excel_book.create_sheet('Summary')
    
    df2 = pd.read_csv(outfile2,sep='\t',lineterminator='\n',header=None)
    rows = dataframe_to_rows(df2,index=False,header=False)
    ws = excel_book['Summary']
    for r_idx, row in enumerate(rows, 1) :
        for c_idx, value in enumerate(row,1) :
            ws.cell(row=r_idx, column=c_idx,value=value)
    os.remove(outfile2)            
    excel_book.save(outfilexlsx)
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - LFRR0011", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def retrt010(f_name):

    escape = '\u001B'
    carriage = '\r'
    colon = '\ua789'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "RETRT010 - RETRO RATE CHANGE APPROVED VOUCHERS REPORT [" + os.path.splitext(f_name[25:35])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    out_str = ("Voucher #\t" + "Vendor\t" + "Vendor Name\t" + "Placement Vendor ID\t" + "Remitance (Payment Amount of new Retro Rate Voucher line)\t" + 
        "Line #\t" + "Case #\t" + "Recipient\t" + "Dob\t" + "New LOD (Level of Difficulty)\t" + "New Program Type (PG)\t" +
        "New Service Type\t" + "New Service Period From\t" + "New Service Period To\t" + "New POS Type Pass-Through Rate (PASS)\t" +
        "New Clothing Rate\t" + "New Administrative Rate\t" + "New Total Rate\t" + "Original Voucher\t" +
        "Check #\t" + "Check Date\t" + "Auth #\t" + "Original LOD (Level of Difficulty)\t" + "Original Program Type (PG)\t" +
        "Original Service Type\t" + "Original Service Period From\t" + "Original Service Period To\t" + "Original POS Type Pass-Through Rate (PASS)\t" + 
        "Original Clothing Rate\t" + "Original Administrative Rate\t" + "Original Total Rate")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        #line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('REPORT DATE') and not line[40:45] == ('RETRO')
            and not line.startswith('BICS') and not line.startswith('VOUCHER AMOUNT')
            and not line.startswith('LINE')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            #if countLines == 1  :
            if line.startswith('VOUCHER NO')  :
                voucherNumber = record[14:22].strip() + '\t'
                vendor = record[32:40].strip() + '\t'
                vendorName = record[55:84].strip() + '\t'
                placementVendorId = record[103:111].strip() + '\t'

            #if countLines == 2  :
            else  :
                if record[6:36].strip() != ''  :
                    if record[0:1].strip() == '' :
                        lineNumber = record[1:5].strip() + '\t'
                        caseNumber = record[6:16].strip() + '\t'
                        recipientId = record[17:25].strip() + '\t'
                        dob = record[26:34].strip() + '\t'
                        newLd = record[35:37].strip() + '\t'
                        newPg = record[38:40].strip() + '\t'
                        newSt = record[41:43].strip() + '\t'
                        newServicePeriodFrom = record[44:52].strip() + '\t'
                        newServicePeriodTo = record[53:61].strip() + '\t'
                        newPass = record[93:100].strip() + '\t'
                        newCloth = record[101:108].strip() + '\t'
                        newAdmin = record[109:116].strip() + '\t'
                        newRate = record[117:124].strip() + '\t'
                        newRemit = record[125:132].strip() + '\t'
                    else :    
                        origVoucher = record[17:25].strip() + '\t'
                        origLd = record[35:37].strip() + '\t'
                        origPg = record[38:40].strip() + '\t'
                        origSt = record[41:43].strip() + '\t'
                        origServicePeriodFrom = record[44:52].strip() + '\t'
                        origServicePeriodTo = record[53:61].strip() + '\t'
                        
                        checkNumber = record[62:69].strip() + '\t'
                        checkDate = record[75:83].strip() + '\t'
                        authNumber = record[84:92].strip() + '\t'
                        
                        origPass = record[93:100].strip() + '\t'
                        origCloth = record[101:108].strip() + '\t'
                        origAdmin = record[109:116].strip() + '\t'
                        origRate = record[117:124].strip()

                        out_str = (voucherNumber + vendor + vendorName + placementVendorId + newRemit + lineNumber +
                            caseNumber + recipientId + dob + newLd + newPg + newSt + newServicePeriodFrom + newServicePeriodTo + 
                            newPass + newCloth + newAdmin + newRate + origVoucher + checkNumber + checkDate + authNumber +
                            origLd + origPg + origSt + origServicePeriodFrom + origServicePeriodTo + origPass + origCloth +
                            origAdmin + origRate)

    # write to results file                                
                        if out_str != ' ':
                            f = open(outfile,"a")
                            f.write(out_str + "\n")
                            f.close()
                            countWrites = countWrites +1 
                            countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - RETRT010", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc1047(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "WINR1047 AUTH OF REPLACEMENT SNAP - EXCEPTION [" + os.path.splitext(f_name[28:39])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("Office\t" + "Unit\t" + "Worker\t" + "Case Number\t" + "Case Name\t" + "Case Type\t" + "Exception Reason")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[14:17].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*') 
            and not line.startswith('ERIE AUTHORIZATION') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[13:45].strip() + '\t'
#                if '/PAYEE' in caseName :
#                    caseName = caseName.replace('/PAYEE','')
#                    payee = 'TRUE\t'
#                else :
#                    payee = '\t'
                caseType = line[46:48].strip() + '\t'
                exceptionReason = line[68:].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + caseType + exceptionReason)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC1025/1040", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def pabs4002(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0

    #pd.options.display.float_format = '{:,.2f}'.format
    #pd.set_option('display.precision', 2)

    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "PABS4002 DPESROLL ELECTRONIC DIR PAYMENT ROLL (RUN TYPE- SINGLE) [" + os.path.splitext(f_name[29:39])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("Category\t" + "Case #\t" + "Case Name\t" + "Payee\t" + "Claim Authorization\t" + "Case Type Code\t" + "Case Type Dec\t" +
               "Issuance Authorization #\t" + "Method of Payment Code\t" + "Method of Payment Desc\t" + "ABEL Budget Recoupment\t" +
               "Check/Benefit #\t" + "ABEL Budget Restrict Amount\t" + "Check/Benefit Amount\t" + "Pay Type\t" + "Pay Type Desc\t" +
               "Local Action Code\t" + "Local Action Desc\t" + "Special Claiming Category\t" + "Supplemental Security Income\t" +
               "Pick-Up Code\t" + "Pick-Up Desc\t" + "HEAP/Special Category (HEAP/SCAT)\t" + "Count of Active Individuals\t" + 
               "State/Federal Charge Code\t" + "State/Federal Charge Desc\t" + "Count of Ind. Designated for S/F\t" + "State/Federal Charge Code2\t" +
               "State/Federal Charge Desc2\t" + "Count of Ind. Designated for S/F2\t" + "State/Federal Charge Code3\t" + 
               "State/Federal Charge Desc3\t" + "Count of Ind. Designated for S/F3")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('CATEGORY') :
            category = record[10:18].strip() + '\t'

        if (not line.startswith('PA') and not line.startswith('REPORT DATE') and not line.startswith('ELECTRONIC')
            and not line.startswith('PICS REPORT') and not line.startswith('BENEFITS') and not line.startswith('CATEGORY') 
            and not line.startswith('CASE NUMBER') and not line.startswith('CS TYPE') and not line.startswith('TOTAL')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0) and record[:60].strip() != '' :

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = record[:11].strip() + '\t'
                caseName = record[15:44].strip() + '\t'
                if '/PAYEE' in caseName or '/ PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    caseName = caseName.replace('/ PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                claimAuthorization =  '\t'
                issuanceAuthorization = record[55:63].strip() + '\t'
                benefitNumber = record[66:78].strip() + '\t'
                checkAmount = record[80:89].strip().replace('$','') + '\t'
                payType = record[91:93].strip()
                payType_dict = {
                    'A1':'Transportation Billed Separately',
                    'A2':'Personal Care Services',
                    'A3':'Personal Care Services - Non-Family Planning',
                    'A4':'Grant Assistance to Guide Dogs',
                    'A6':'Approved Facility/Congregate Care Facility',
                    'A7':'Refrigerator Rental Allowance for Homeless Family Temporarily Placed in a Hotel/Motel',
                    'C0':'Replacement of Stolen Cash (EAA)',
                    'C1':'Replacement of Lost/Mismanaged Cash (EAA)',
                    'C2':'Furniture Allowance for Establishment of a Home',
                    'C3':'Payment for Services and Supplies Received Prior to EAA Application (EAA)',
                    'C4':'Payment for Services and Supplies Received Prior to SSI Application (EAA)',
                    'C5':'Nutritional Requirements (EAA)',
                    'C7':'Transportation for Homeless Families',
                    'C9':'Payment for Services to Cope with an Emergency',
                    'D1':'IV-D Payment',
                    'D2':'Child Visitation Allowance',
                    'D3':'Excess Current Support',
                    'D4':'Excess Support Arrears',
                    'D7':'Transitional Services Payments (TANF Ineligible Due to Employment) (CT 11, 12, 16, 17, 19)',
                    'D8':'Private Adult Care Institution',
                    'D9':'Diversion Transportation Payment (CT 11, 12, 19 Only)',
                    'E1':'Grant to Essential Person (PA)',
                    'E3':'Rental Supplement (RENT-SUP)',    
                    'E5':'Enhanced Shelter Allowance',
                    'E6':'Fuel for Heating Refund',
                    'E7':'Electricity',
                    'E8':'Emergency Shelter Payment',
                    'E9':'Restricted PNA',
                    'F2':'Food Stamp Employment/Training Related Expenses',
                    'F3':'FSE & T Dependent Care Payment',
                    'F5':'Shelter Related Expenses and Mortgage/Tax Arrears',
                    'F6':'Diversion Rental Payment (CT 11, 12, 19 Only)',
                    'G1':'Shelter/R&B to Guardian (CT 12 & 17 Only)',
                    'G5':'Karmalla',
                    'G8':'COLAJ',
                    'H0':'Heating Equip. Repair/Replacement Estimates',
                    'H1':'HEAP Regular Benefit Heater',
                    'H2':'HEAP Cooling',
                    'H5':'HEAP Emergency Benefit - Repair Heating Equipment',
                    'H6':'HEAP Emergency Benefit - Shelter/Relocation',
                    'H7':'HEAP Emergency Benefit - Replace Heating Equipment',
                    'H8':'HEAP Emergency Benefit - Propane Tank Deposit',
                    'H9':'HEAP Supplemental Benefit',
                    'HX':'HEAP Regular Benefit Heat Included',
                    'J1':'HEAP Emergency Benefit - Additional Benefit',
                    'J2':'HEAP Reissue Benefit',
                    'J3':'HEAP Clean and Tune',
                    'J9':'HEAP Additional Regular Benefit',
                    'K1':'CAP Grant',
                    'K3':'CAP Support Reconciliation',
                    'L1':'Drug and Sickroom Supplies',
                    'L2':'Prosthetic Appliances and Eye Glasses',
                    'L3':'Lab and Radiologist Services Billed Separately',
                    'L4':'Health Insurance Continuation - 185% Poverty',
                    'L5':'Health Insurance Continuation - 100% Poverty',
                    'L6':'Health Insurance Continuation - Expedited Payment',
                    'M1':'Hospital Services - Inpatient - Public',
                    'M2':'Hospital Services - Inpatient - Private',
                    'M3':'Hospital Services - Outpatient - Public',
                    'M4':'Hospital Services - Outpatient - Private',
                    'M5':'Skilled Nursing Home - Private',
                    'M6':'Skilled Nursing Home - Public',
                    'M7':'Health Related Facilities - Public - Developmental Disabilities',
                    'M8':'Health Related Facilities - Public - Other',
                    'M9':'Health Related Facilities - Private - Developmental Disabilities',
                    'NX':'Nominal HEAP Benefit',
                    'N1':'Emergency Non-Utility Fuel Payment',
                    'N2':'Child Support Due Client - Period of Ineligibility (CT 11, 12, 16, 17 Only)',
                    'P1':'Health Related Facilities - Private - Other',
                    'P2':'Free Standing Clinics',
                    'P3':'Physician''s Services',
                    'P4':'Dental Services',
                    'P5':'Other Practitioner''s Services',  
                    'P6':'Child Caring Agencies Per Diem Costs',
                    'P7':'Home Health Aide''s Services',
                    'P8':'Nursing Services In-Home',
                    'P9':'Care at Home',
                    'Q1':'Family Shelter Tier I',
                    'Q2':'Family Shelter Tier II',
                    'Q4':'Transitional Housing',
                    'Q5':'Upfront Cash Security Deposit',
                    'Q6':'Residential Domestic Violence',
                    'R0':'Legally Exempt In-Home Child Care Relative (Full Time)',
                    'R1':'Legally Exempt In-Home Child Care Relative (Part Time)',
                    'R2':'Legally Exempt Family Child Care Relative (Full Time)',
                    'R3':'Legally Exempt Family Child Care Relative (Part Time)',
                    'R4':'Legally Exempt Family Child Care Non-Relative (Full Time)',
                    'R5':'Legally Exempt Family Child Care Non-Relative (Part Time)',
                    'R6':'School Age Child Care Program - Part Time',
                    'R7':'Transportation',
                    'R8':'School Age Child Care Program - Full Time',
                    'R9':'Employment and Training Essential Needs',
                    'T1':'On the Job Training Grant',
                    'T2':'Extended Supportive Services',
                    'T3':'Training Tuition and Fees',
                    'T5':'TANF Services Block Grant/Flexible Funding',
                    'U1':'FHP PAP Premium (Disabled 6/19/22)',
                    'U2':'FHP PAP Deductible (Disabled 6/19/22)',
                    'U3':'FHP Co-Pay Differential (Disabled 6/19/22)',
                    'U4':'FHP PAP Other (Disabled 6/19/22)',
                    'U5':'FHP PAP Co-Insurance (Disabled 6/19/22)',
                    'W1':'Court Ordered Retroactive Payment - Check',
                    'W2':'Court Ordered Retroactive Payment - Cash',
                    '02':'Legally Exempt Group Child Care (Full Time)',
                    '03':'Legally Exempt Group Child Care (Part Time)',
                    '04':'HEAP Emergency Benefit - Non-Utility',
                    '05':'Case Recurring Grant',
                    '06':'Partial Allowance',
                    '07':'Underpayment Adjustment',
                    '10':'Shelter',
                    '11':'Fuel',
                    '12':'Utilities',
                    '13':'Guaranteed Utility Account Payment',
                    '14':'TEAP (Transitional Employment Advancement Program)',
                    '16':'HEAP Emergency Benefit - Domestic Heat-related Utility',
                    '17':'HEAP Emergency Benefit - Utility',
                    '18':'Child Support Disregard',
                    '19':'Health Maintenance Organization Co-Payment',
                    '24':'Health Insurance Premiums',
                    '25':'Life Insurance Premiums',
                    '26':'Chattel Mortgage',
                    '28':'Cooking Fuel',
                    '29':'Related Foster Care Expenses',
                    '30':'Legally Exempt In-Home Child Care Non-Relative (Full Time)',
                    '31':'Legally Exempt In-Home Child Care Non-Relative (Part Time)',
                    '32':'Day Care Family Home (Full-Time)',
                    '33':'Day Care Family Home (Part-Time)',
                    '34':'Day Care Group Family (Full-Time)',
                    '35':'Home Delivered Meals (PA Grant)',
                    '36':'Day Care Group Family (Part-Time)',
                    '37':'Day Care Center (Full-Time)',
                    '38':'Day Care Center (Part-Time)',
                    '39':'Disaster Card Issuance',
                    '40':'Room and Board',
                    '41':'Appliance Repair Cost Estimate',
                    '42':'Replacement of Lost/Stolen Cash Grant',
                    '43':'Heating Equipment - Repairs/Replacement',
                    '44':'Cooking Stove - Repairs/Replacement',
                    '45':'Refrigerator - Repairs/Replacement',
                    '46':'Private Rent',
                    '47':'Mortgages, Taxes and Assessments on Client Owned Homes',
                    '48':'Public Housing Rent Allowance',
                    '49':'Housing Development Cooperative Unit',
                    '50':'Temporary Residence in Hotels & Motels',
                    '51':'Cost of Repairs to Recipient Owned Home',
                    '52':'Emergency Allowance to Forestall Eviction/Foreclosure',
                    '54':'Restaurant Allowance - Dinner',
                    '55':'Restaurant Allowance - Lunch and Dinner',
                    '56':'Restaurant Allowance - All Meals',
                    '57':'Emergency Food Grant Allowance',
                    '58':'Natural Gas',
                    '59':'Other Than Natural Gas',
                    '60':'Emergency Utility Payment to Prevent Shut-Off or Restore Service',
                    '62':'Taxes and Interest',
                    '63':'Water Bills',
                    '64':'Real Property Expenses',
                    '65':'Moving Expenses',
                    '66':'Storing Expenses',
                    '67':'Payment on Security Agreement',
                    '68':'Broker''s Finder''s Fee',
                    '69':'Cash Grant Unrestricted',
                    '70':'Cash Replacement for Lost, Stolen or Mismanaged SSI Grant',
                    '71':'Other',
                    '72':'Living Room',
                    '73':'Bedroom with Single Bed',
                    '74':'Bedroom with Two Single Beds',
                    '75':'Bedroom with Double Bed',
                    '76':'Kitchen (Excluding Appliances)',
                    '77':'Range',
                    '78':'Refrigerator',
                    '79':'Bathroom',
                    '81':'Housekeeping',
                    '82':'Camp Fees',
                    '83':'Red Cross',
                    '84':'Payment on Furniture',
                    '85':'Burial',
                    '86':'Transportation Expenses - Removal from State',
                    '87':'Cash Grant for Income Not Received',
                    '88':'Cabinet for Linens',
                    '89':'Stove for Heating',
                    '90':'Cost of Clothing',
                    '91':'Expedited ATP',
                    '92':'Food Stamp Replacement',
                    '93':'Single Issuance ATP (Not Expedited)',
                    '94':'FS Retroactive Benefits',
                    '95':'FS Restored Benefits',
                    '96':'FS Ongoing Benefits',
                    '97':'Food Stamps Supplemental',
                    '98':'Home Repairs (PA)'
                }
                payTypeDesc = payType_dict[payType] + '\t'
                payType = payType + '\t'
                specialClaiming = record[95:97].strip() + '\t'
                pickupCode = record[103:104].strip()
                pickupCode_dict = {'1':'Mailed', '2':'NYSES', '3':'Agency Pick-Up', '4':'Vendor Direct', '5':'Delivered by Agency', '6':'Other'}
                pickupCodeDesc = pickupCode_dict[pickupCode] + '\t'
                pickupCode = pickupCode + '\t'
                activeIndividuals = record[107:109].strip() + '\t'
                stateFederalChargeCode_dict = {
                    '03':'American Repatriate',
                    '04':'NativeAmerican on NYS Reservation',
                    '05':'OMH/OPWDD Release',
                    '07':'OMH/OPWDD Inpatient',
                    '08':'OMH/OPWDD Family Care',
                    '11':'Oxford Home Resident',
                    '18':'State-Operated ICF',
                    '19':'Privately-Operated ICF',
                    '21':'VORCCA (Voluntary-Operated Residential Care Center for Adults - Non-621)',
                    '22':'SOCR (State-Operated Community Residence- Non-621)',
                    '23':'VOFC (Voluntary-Operated Family Care), OMH Home & Community Based Services (HCBS) Waiver',
                    '24':'VOCR (Voluntary-Operated Community Residence- Non-621)',
                    '25':'VOCR (Voluntary-Operated Community Residence - 621)',
                    '26':'SOCR (State-Operated Community Residence- KEYES) [OMH, OPWDD Only]',
                    '27':'SOCR (State-Operated Community Residence- Non-KEYES) [OMH, OPWDD Only]',
                    '28':'SORCCA (State-Operated Residential Care Center for Adults) [OMH Only]',
                    '29':'VORCCA (Voluntary-Operated Residential Care Center for Adults)',
                    '30':'Refugee Assistance Programs (RCA/RMA)',
                    '31':'Unaccompanied Refugee Minor',
                    '37':'Relocated Relative of an Institutionalized Veteran',
                    '50':'Home Care-State Charge (Case Type 20 Only)',
                    '60':'TANF IneligibleAlien',
                    '63':'TANF Individual Exceeding 5 Year Limit',
                    '64':'TANF NativeAmerican on NYS Reservation Exceeding 5 Year Limit',
                    '65':'PRUCOL Pregnant FFP',
                    '66':'PRUCOL under 21 FFP',
                    '67':'PRUCOL Non-Qualified Alien',
                    '68':'QualifiedAlien Not MOE Eligible',
                    '70':'Incarcerated Local Jail',
                    '71':'Incarcerated in a NYS Department of Corrections and Community Supervision correctional facility',
                    '72':'Pregnant Consumer with ACI=E 88 State Charge/Federal Charge Expired'
                }
                stateFederalChargeCode = record[113:115].strip()
                if stateFederalChargeCode != '' :
                    stateFederalChargeDesc = stateFederalChargeCode_dict[stateFederalChargeCode] + '\t'
                    stateFederalChargeCode = stateFederalChargeCode + '\t'
                    stateFederalChargeNumber = record[116:118].strip() + '\t'
                else :
                    stateFederalChargeCode = '\t'
                    stateFederalChargeDesc = '\t'
                    stateFederalChargeNumber = '\t'
                stateFederalChargeCode2 = record[119:121].strip()
                if stateFederalChargeCode2 != '' :
                    stateFederalChargeDesc2 = stateFederalChargeCode_dict[stateFederalChargeCode2] + '\t'
                    stateFederalChargeCode2 = stateFederalChargeCode2 + '\t'
                    stateFederalChargeNumber2 = record[122:124].strip() + '\t'
                else :
                    stateFederalChargeCode2 = '\t'
                    stateFederalChargeDesc2 = '\t'
                    stateFederalChargeNumber2 = '\t'
                stateFederalChargeCode3 = record[125:127].strip()
                if stateFederalChargeCode3 != '' :
                    stateFederalChargeDesc3 = stateFederalChargeCode_dict[stateFederalChargeCode3] + '\t'
                    stateFederalChargeCode3 = stateFederalChargeCode3 + '\t'
                    stateFederalChargeNumber3 = record[128:130].strip()
                else :
                    stateFederalChargeCode3 = '\t'
                    stateFederalChargeDesc3 = '\t'
                    stateFederalChargeNumber3 = ' '

            if countLines == 2  :
                caseType = record[45:51].strip()
                caseType_dict = {'FA':'11', 'SN-FP':'12', 'ADC-FC':'13', 'SN-CSH':'16', 'SN-FNP':'17', 'EAA':'18',
                                 'EAF':'19', 'MA':'20', 'MA-SSI':'22', 'FHP':'24', 'NPA-FS':'31', 'FS-MIX':'32', 'HEAP':'60'}
                caseTypeDesc = caseType_dict[caseType] + '\t'
                caseType = caseType + '\t'
                methodOfPayment = record[54:56].strip()
                methodOfPayment_dict = {'01':'Unrestricted', '02':'Vendor as Authorized', '03':'Vendor as Billed',
                                        '04':'Vendor as Billed Subject to Limit', '05':'Associated Name A',
                                        '06':'Associated Name B', '07':'Vendor as Billed Subject to Review',
                                        '08':'Other', '09':'Restricted', '10':'Food Stamp Cash Out',
                                        '11':'Vendor Line of Credit - HEAP', '12':'Previously Issued Emergency Card'}
                methodOfPaymentDesc = methodOfPayment_dict[methodOfPayment] + '\t'
                methodOfPayment = methodOfPayment + '\t'
                recoupmentAmount = record[58:65].strip().replace('$','') + '\t'
                restrictAmount = record[66:73].strip().replace('$','') + '\t'
                localAction = record[91:92].strip()
                localAction_dict = {'1':'Check/ATP Issued', '2':'Prepare and Issue Check/ATP', '3':'Hold',
                                    '4':'Release', '5':'Cancel', '6':'Other', '7':'Replacement Check/ATP Issued',
                                    '8':'Prepare and Issue Replacement Check/ATP', '9':'Void'}
                localActionDesc = localAction_dict[localAction] + '\t'
                localAction = localAction + '\t'
                supplementalSecurityIncome = '\t'
                heapScat = '\t'

                out_str = (category + caseNumber + caseName + payee + claimAuthorization + caseType + caseTypeDesc + 
                           issuanceAuthorization + methodOfPayment + methodOfPaymentDesc + recoupmentAmount + benefitNumber +
                           restrictAmount + checkAmount + payType + payTypeDesc + localAction + localActionDesc +
                           specialClaiming + supplementalSecurityIncome + pickupCode + pickupCodeDesc + heapScat + activeIndividuals +
                           stateFederalChargeCode + stateFederalChargeDesc + stateFederalChargeNumber + stateFederalChargeCode2 + 
                           stateFederalChargeDesc2 + stateFederalChargeNumber2 + stateFederalChargeCode3 + 
                           stateFederalChargeDesc3 + stateFederalChargeNumber3)  

    # write to results file                                
                if out_str != ' ':
#                    print("f_name is " + f_name + "and Category is " + line)
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=0)
    df['Check/Benefit #'] = df['Check/Benefit #'].apply(str)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=True)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - PABS4002", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    # Create message object
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = recipient
    msg['Subject'] = 'Email with file attachment'
    # Add body text to message
    body_text = 'Please see attached file.'
    msg.attach(MIMEText(body_text))
    # Add file attachment to message
    file_path = outfilexlsx
    with open(file_path, 'rb') as f:
        attachment = MIMEApplication(f.read(), _subtype='xlsx')
        attachment.add_header('Content-Disposition', 'attachment', filename=outfilexlsx)
        msg.attach(attachment)
    # Send the email using an SMTP server
    smtp_server = 'smtp.erie.gov'
    smtp_port = 587
    smtp_username = 'sulkowsf'
    smtp_password = 'Peytonfrog-328'
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)

    return

#===================================================================================================

def cns00125(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0

    outfile = "CNS00125 WORKER NAME & TELEPHONE [" + os.path.splitext(f_name[35:45])[0] + "].csv"
    if os.path.exists(outfile):
        os.remove(outfile)
    errfile = "CNS00125 WORKER NAME & TELEPHONE-error [" + os.path.splitext(f_name[35:45])[0] + "].csv"
    if os.path.exists(errfile):
        os.remove(errfile)


    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("Office," + "Unit," + "Worker," + "Worker Name," + "Telephone Number")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
        f = open(errfile,"a")
        f.write(out_str + "\n")
        f.close()
    

    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('60') and not line.startswith('REPORT') 
            and not line.startswith('*') and not line.startswith('WORKER') 
            and not line.startswith('PERIOD') and not line.startswith('DISTRICT')
            and not line.startswith('PROGRAM') and not line.startswith('-------') 
            and not line.startswith('END OF REPORT')  
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) >= 48):

            countLines = countLines + 1

            if countLines == 1  :
                office = record[25:28].strip() + ','
                unit = record[34:39].strip() + ','
                worker = record[45:50].strip() + ','
                workerName = record[59:93].strip().replace(',',' ') + ','
                telephoneNumber = record[96:108].strip()

                out_str = (office + unit + worker + workerName + telephoneNumber)

    # write to results file  
                if len(office) == 4 and len(unit) == 6 and len(worker) == 6:
                    if out_str != ' ':
                        f = open(outfile,"a")
                        f.write(out_str + "\n")
                        f.close()
                        countWrites = countWrites +1 
                        countLines = 0
                else :
                    f = open(errfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    
              
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - CNS00125", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================


main()

