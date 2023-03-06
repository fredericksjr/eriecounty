#!/usr/bin/env python

"""convertNASReports.py: read in heap autopay txt file file and export as single record text file (outAutopay.txt) """

__author__   = "Frederick Sulkowski"
__email__   = "frederick.sulkowski@erie.gov"


from argparse import ArgumentDefaultsHelpFormatter
from ftplib import ftpcp
from turtle import goto
import os
import pandas as pd
from tkinter import messagebox



def main():

    # read files in current directory
    for f_name in os.listdir('.'):
        if (f_name.startswith('001-X-WMMRELNA-14') and f_name.endswith('.txt')) or (f_name.startswith('001-X-WMMRRJNA-14') and f_name.endswith('.txt')) :
            wmmrelna(f_name)

        if f_name.startswith('001-X-WMMREXNA-14') and f_name.endswith('.txt') :
            wmmrexna(f_name)
        if f_name.startswith('001-X-WM4DELNA-14') and f_name.endswith('.txt') :
            wm4delna(f_name)
        if f_name.startswith('001-X-WM4DEXNA-14') and f_name.endswith('.txt') :
            wm4dexna(f_name)
        if f_name.startswith('001-X-14-WMSCSDX5') and f_name.endswith('.txt') :
            wmscsdx5(f_name)
        if (f_name.startswith('001-X-14-WMSBF630') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSBP630') and f_name.endswith('.txt')) :
            wmsbf630(f_name)
        if (f_name.startswith('001-X-14-WMSBF650') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSBP650') and f_name.endswith('.txt')) :
            wmsbf650(f_name)
        if (f_name.startswith('001-X-14-WMSBF650') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSBP650') and f_name.endswith('.txt')) :
            wmsbf650(f_name)
        if (f_name.startswith('001-X-14-WMSA5257') and f_name.endswith('.txt')) :
            wmsa5257(f_name)
        if (f_name.startswith('001-X-14-WMSC1025') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1040') and f_name.endswith('.txt')) :
            wmsc1025(f_name)
        if (f_name.startswith('001-X-14-WMSC1026') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1041') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1661') and f_name.endswith('.txt')) :
            wmsc1026(f_name)
        if (f_name.startswith('001-X-14-WMSC2027') and f_name.endswith('.txt')) :
            wmsc2027(f_name)            
        if (f_name.startswith('001-X-14-WMSBHIGH') and f_name.endswith('.txt')) :
            wmsbhigh(f_name)            
        if (f_name.startswith('001-X-14-WMSC4210') and f_name.endswith('.txt')) :
            wmsc4210(f_name)            
        if (f_name.startswith('WINR0799_PDF') and f_name.endswith('.txt')) :
            winr0799(f_name)            
        if (f_name.startswith('CSPROD-BBP101PRTF') and f_name.endswith('.txt')) :
            wbb101(f_name)            
        if (f_name.startswith('PERIE-PYCHECKS') and f_name.endswith('.txt')) :
            lfrr0011(f_name)            
        if (f_name.startswith('PERIE-SVRETVCH') and f_name.endswith('.txt')) :            
            retrt010(f_name)
        if (f_name.startswith('001-X-14-WMSC1047') and f_name.endswith('.txt')) or (f_name.startswith('001-X-14-WMSC1662') and f_name.endswith('.txt')):
            wmsc1047(f_name)
            

#===================================================================================================

def wmmrelna(f_name):

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "OLD DEFCT\t" + "NEW DEFCT\t" + "NET CHNG\t" +
                "SPEC ALERT\t" + "OLD ETLMT\t" + "NEW ETLMT\t" + "NET CHNG\t" + "REJ/EXCEPT")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            if 'WMMRRJNA' in f_name:
                office = record[54:59].strip() + '\t'
                unit = record[65:70].strip() + '\t'
                worker = record[80:86].strip() + '\t'
            else:
                office = record[44:49].strip() + '\t'
                unit = record[56:61].strip() + '\t'
                worker = record[72:78].strip() + '\t'

        if (not line.startswith('WMRIQ5') and not line.startswith('DISTRICT') and not line.startswith('OLD DEFCT') 
            and not line.startswith('CASE NAME') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[26:47].strip() + '\t'
                caseNumber = record[48:58].strip() + '\t'
                oldDefct = record[60:67].strip() + '\t'
                newDefct = record[72:78].strip() + '\t'
                netChng1 = record[82:89].strip() + '\t'
                specAlert = record[94:].strip() + '\t'

            if countLines == 2:
                oldEtlmt = record[60:67].strip() + '\t'
                newEtlmt = record[71:78].strip() + '\t'
                netChng2 = record[82:89].strip() + '\t'
                rejExcept = record[94:]
                endPos = rejExcept.find(carriage)
                rejExcept = rejExcept[0:endPos].strip()
                out_str = (office + unit + worker + caseName + caseNumber + oldDefct + newDefct + netChng1 + specAlert + 
                           oldEtlmt + newEtlmt + netChng2 + rejExcept)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    if 'WMMRRJNA' in f_name:
        messagebox.showinfo("Completed - WMMRRJNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + "!")
    else:
        messagebox.showinfo("Completed - WMMRELNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + "!")

    return

#===================================================================================================

def wmmrexna(f_name):

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "EXCEPTED REASON")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            office = record[54:59].strip() + '\t'
            unit = record[65:70].strip() + '\t'
            worker = record[80:86].strip() + '\t'

        if (not line.startswith('WMRIQ4') and not line.startswith('DISTRICT') and not line.startswith('CASE NAME') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[28:55].strip() + '\t'
                caseNumber = record[56:66].strip() + '\t'
                exceptedReason = record[84:].strip()
                out_str = (office + unit + worker + caseName + caseNumber + exceptedReason)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMMREXNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wm4delna(f_name):

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "MESSAGE #1\t" + "PREV 2\t" +
                "PREV 1\t" + "CUR IVD\t" + "NEW FS\t" + "CUR FA\t" + "SPEC ALERTS\t" + "MESSAGE #2\t" + "CUM 2\t" +
                "CUM 1\t" + "CUR CUM\t" + "OLD FS\t" +"PRE FA\t" + "ACTION CODE\t" + "MONTH OBL\t" + "EXMT 2\t" +
                "EXMT 1\t" + "EXMT CUR")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            office = record[47:52].strip() + '\t'
            unit = record[58:64].strip() + '\t'
            worker = record[75:81].strip() + '\t'

        if (not line.startswith('WMRIV2') and not line.startswith('DISTRICT') and not line.startswith('CASE NAME')
            and not line.startswith('CASE NUMBER') and not line.startswith('MONTH OBL') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[9:39].strip() + '\t'
                message1 = record[39:50].strip() + '\t'
                prev2 = record[54:61].strip() + '\t'
                prev1 = record[65:72].strip() + '\t'
                curIvd = record[76:83].strip() + '\t'
                newFs = record[87:94].strip() + '\t'
                curFa = record[98:105].strip() + '\t'
                specAlerts = record[109:].strip() + '\t'
                
            if countLines == 2  :
                caseNumber = record[14:24].strip() + '\t'
                message2 = record[39:50].strip() + '\t'
                cum2 = record[54:61].strip() + '\t'
                cum1 = record[65:72].strip() + '\t'
                curCum = record[76:83].strip() + '\t'
                oldFs = record[87:94].strip() + '\t'
                preFa = record[98:105].strip() + '\t'
                actionCode = record[109:].strip() + '\t'

            if countLines == 3  :
                monthObl = record[39:50].strip() + '\t'
                exmt2 = record[54:61].strip() + '\t'
                exmt1 = record[65:72].strip() + '\t'
                exmtCur = record[76:83].strip()
                out_str = (office + unit + worker + caseName + caseNumber + message1 + prev2 + prev1 + curIvd + newFs +
                           curFa + specAlerts + message2 + cum2 + cum1 + curCum + oldFs + preFa + actionCode + monthObl +
                           exmt2 + exmt1 + exmtCur)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WM4DELNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wm4dexna(f_name):
 
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "WMRIV1 DEXNA MRBA - EXCEPTION LIST FOR IVD [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "MESSAGE #1\t" + "PREV 2\t" +
                "PREV 1\t" + "CUR IVD\t" + "CUR FA\t" + "EXCEPTION REASON\t" + "MESSAGE #2\t" + "CUM 2\t" + "CUM 1\t" +
                "CUR CUM\t" + "PRE FA\t" +"ACTION REQUIRED\t" + "MONTH OBL\t" + "EXMT 2\t" + "EXMT 1\t" + "EXMT CUR")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('FOR MASS REBUDGETING'):
            startDetail = False

        if line.startswith('DISTRICT') and startDetail == True:
            office = record[47:52].strip() + '\t'
            unit = record[58:64].strip() + '\t'
            worker = record[75:81].strip() + '\t'

        if (not line.startswith('WMRIV1') and not line.startswith('DISTRICT') and not line.startswith('CASE NAME')
            and not line.startswith('CASE NUMBER') and not line.startswith('MONTH OBL') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseName = record[9:39].strip() + '\t'
                message1 = record[39:50].strip() + '\t'
                prev2 = record[54:61].strip() + '\t'
                prev1 = record[65:72].strip() + '\t'
                curIvd = record[76:83].strip() + '\t'
                curFa = record[87:93].strip() + '\t'
                exceptionReason = record[98:].replace(nullChar,' ').strip() + '\t'

            if countLines == 2  :
                caseNumber = record[14:24].strip() + '\t'
                message2 = record[39:50].strip() + '\t'
                cum2 = record[54:61].strip() + '\t'
                cum1 = record[65:72].strip() + '\t'
                curCum = record[76:83].strip() + '\t'
                preFa = record[87:93].strip() + '\t'
                actionRequired = record[98:].strip() + '\t'

            if countLines == 3  :
                monthObl = record[39:50].strip() + '\t'
                exmt2 = record[54:61].strip() + '\t'
                exmt1 = record[65:72].strip() + '\t'
                exmtCur = record[76:83].strip()
                out_str = (office + unit + worker + caseName + caseNumber + message1 + prev2 + prev1 + curIvd +
                           curFa + exceptionReason + message2 + cum2 + cum1 + curCum + preFa + actionRequired +
                           monthObl + exmt2 + exmt1 + exmtCur)

            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
               
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)               
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")
    os.remove(outfile)            

    messagebox.showinfo("Completed - WM4DEXNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmscsdx5(f_name):
    
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0
    statusNext = ' '

    outfile = "ec_" + f_name
    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfile):
        os.remove(outfile)
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("STATUS\t" + "CASE TYPE\t" + "SSN\t" + "LAST NAME\t" + "FIRST NAME\t" + "M I\t" + "CASE NUMBER\t" +
                "CIN\t" + "OLD ETMNT\t" + "NEW ETMNT\t" + "NEW ALMNT\t" + "F T\t" + "AUTH FROM\t" + "AUTH TO\t" + "TRANSACTION DISPOSITION\t" +
                "TRANSACTION DISPOSITION\t" + "TRANSACTION DISPOSITION\t" + "TRANSACTION DISPOSITION")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if statusNext == 'Y' and not line.startswith(escape):
            status = record[0:64].strip() + '\t'
            statusNext = ' '

        if line.startswith('TYPE'):
            statusNext = 'Y'
        
        if (not line.startswith('DISTRICT') and not line.startswith('AUTO SDX') and not line.startswith('AUTOMATED')
            and not line.startswith('TRANSACTION CONTROL REPORT') and not line.startswith('CASE') 
            and not line.startswith('TYPE') and not line.startswith('EXCEPTION') and not line.startswith('ELIGIBLE')
            and not line.startswith('RFC') and not line.startswith('NON WORKER') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0) :

            if countLines == 3 :
                countLines = 0

            if countLines == 2 :
                recordCheck = record[0:20].strip()

                if recordCheck == "" :
                    transactionDisposition3 = record[26:60].strip() + '\t'
                    transactionDisposition4 = record[61:95].strip()
                else :
                    transactionDisposition3 = ' ' + '\t'
                    transactionDisposition4 = ' '
                    countLines = 0

                out_str = (status + caseType + ssn + lastName + firstName + mi + caseNumber + cin + oldEtmnt +
                    newEtmnt + newAlmnt + ft + authFrom + authTo + transactionDisposition1 + transactionDisposition2 +
                    transactionDisposition3 + transactionDisposition4) 
                           
            # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 

            countLines = countLines + 1

            if countLines == 1 :
                caseType = record[0:4].strip() + '\t'
                if record[7:8].strip() == 'd' :
                    record2 = '       d           ' + record[11:]
                    record = record2
                ssn = record[7:16].strip() + '\t'
                lastName = record[19:39].strip() + '\t'
                firstName = record[39:50].strip() + '\t'
                mi = record[50:51].strip() + '\t'
                caseNumber = record[54:64].strip() + '\t'
                cin = record[67:75].replace(nullChar,' ').strip() + '\t'
                oldEtmnt = record[78:86].strip() + '\t' 
                newEtmnt = record[89:97].strip() + '\t'
                newAlmnt = record[100:107].strip() + '\t'
                ft = record[110:111].strip() + '\t'
                authFrom = record[114:122].strip() + '\t'
                authTo  = record[122:130].strip() + '\t'

            if countLines == 2 :
                transactionDisposition1 = record[26:60].strip() + '\t'
                transactionDisposition2 = record[61:95].strip() + '\t'

    out_str = (status + caseType + ssn + lastName + firstName + mi + caseNumber + cin + oldEtmnt +
        newEtmnt + newAlmnt + ft + authFrom + authTo + transactionDisposition1 + transactionDisposition2 +
        transactionDisposition3 + transactionDisposition4)
                           
    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
        countWrites = countWrites +1 

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WM4DEXNA", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsbf630(f_name):
    
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0
    heapPerson = ' '

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("CASE NUMBER\t" + "CASE NAME\t" + "LASTNAME\t" + "FIRSTNAME\t" + "MI\t" + "EXCEPTION REASON")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('DISTRICT') and not line.startswith('LIST OF')
            and not line.startswith('CASE NUMBER') and not line.startswith(escape)) :
            # and not line.startswith(carriage) and len(line) != 0) :

                    countLines = countLines + 1
                    lengthOfLine = len(line)

            # format 3 lines of data into one            
                    if line[0] != '\"':
                        line = ("\"" + line + "\"")

                    heapPerson = heapPerson + line

                    if countLines == 1 and lengthOfLine < 118:
                        heapPerson = heapPerson[:116] + "   \""

                    if countLines == 1:
                        new_str = heapPerson.replace("\"", " ")

            # case number (caseNumber)
                        caseNumber = new_str[2:13].strip() + '\t'
            # case name (caseName)                
                        if "." in new_str[15:51]:
                            new_str = new_str[:15] + new_str[15:51].replace("."," ") + new_str[51:]

                        if ", " in new_str[15:51] or " ," in new_str[15:51]:
                            new_str = new_str.replace(", ",",")
                            new_str = new_str.replace(" ,",",")
                    
                        tempName = new_str[15:50].strip()
                        caseName = new_str[15:50].strip() + '\t'


#                        def get_name_parts(name):
                        comma_split = tempName.split(',')
                        tempLast = comma_split[0]
                        caseLast = str(tempLast) + '\t'

                        try:
                            tempFirst = comma_split[1].split(',')[0]
                            caseFirst = str(tempFirst) + '\t'
                        except IndexError:
                            caseFirst = ' \t'

                        try:
                            tempMI = comma_split[2].split(',')[0]
                            caseMI = str(tempMI) + '\t'
                        except IndexError:
                            caseMI = ' \t'

            # exception reason (exceptionReason)
                        exceptionReason = new_str[52:76].strip()


               
                        out_str = caseNumber + caseName + caseLast + caseFirst + caseMI + exceptionReason

            # write to results file                                
                        if out_str != ' ':
                            f = open(outfile,"a")
                            f.write(out_str + "\n")
                            f.close()

                            countWrites = countWrites +1 
                            countLines = 0
                            heapPerson = ' '

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None,low_memory=False)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSBF630", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsbf650(f_name):
    
    escape = '\u001B'
    carriage = '\r'
    nullChar = '\u0000' 

    # setup count variables
    countLines = 0
    countWrites = 0

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    # read file
    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("CASE NUMBER\t" + "CASE NAME\t" + "ADDRESS\t" + "CITY\t" + "STATE\t" + "ZIP\t" + "Local Office\t" + "UNIT\t" + 
                            "WORKER\t" + "ACT (Local Action Code)\t" + "PAY TYPE\t" + "Method\t" + "AMOUNT\t" + "IS (Issuance Code)\t" + 
                            "SC (Schedule Code)\t" + "PU (Pick Up Code)\t" + "PERIOD FROM\t" + "PERIOD TO\t" + "FUEL TYPE\t" +
                            "SPC CLM (Special Claim\t)" + "LOCAL USAGE\t" + "VP (Voucher Payment\t" + "HH SIZE\t" + "INCOME\t" +
                            "HEAP VENDOR ID\t" + "CUSTOMER ACCT NO\t" + "NX (Heap Nominal Benefit\t" + "CASE TYPE")

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

        # setup variables
        heapPerson = ' '
    
        for line in lines:
            line = line.strip().upper()
            line = line.strip()

            if (not line.startswith('CASE NAME') and not line.startswith('DISTRICT') and not line.startswith('ADDRESS') 
                and not line.startswith('CASE NUMBER') and not line.startswith('RC132') and not line.startswith('* * *  DIS') 
                and not line.startswith('"*  TOTAL') and not line.startswith('*  TOTAL') and not line.startswith("\u001B")):

                countLines = countLines + 1
                lengthOfLine = len(line)

                # format 3 lines of data into one            
                if line[0] != '\"':
                    line = ("\"" + line + "\"")

                heapPerson = heapPerson + line

                if countLines == 1 and lengthOfLine < 118:
                    heapPerson = heapPerson[:116] + "   \""

    #            if countLines == 2 and heapPerson[158] == ' ':
    #                heapPerson = heapPerson[:158] + heapPerson[160:]

                if countLines == 3 and heapPerson[188] == '"':
                    heapPerson = heapPerson[:189] + '           ' + heapPerson[189:]

                if countLines == 3:
                    new_str = heapPerson.replace("\"", " ")
                    
                    caseNumber = new_str[2:14].strip() + '\t'
                    caseName = new_str[15:51].strip() + '\t'
                    caseAddress = new_str[121:157].strip() + '\t'
                    caseCity = new_str[201:217].strip() + '\t'
                    caseState = new_str[219:221].strip() + '\t'
                    caseZip = new_str[225:230].strip() + '\t'
                    localOffice = new_str[52:55].strip() + '\t'
                    caseUnit = new_str[158:163].strip() + '\t'
                    caseWorker = new_str[238:243].strip() + '\t'
                    localActionCode = new_str[60:61].strip() + '\t'
                    payType = new_str[64:66].strip() + '\t'
                    caseMethod = new_str[70:72].strip() + '\t'
                    caseAmount = new_str[73:82].strip() + '\t'
                    issuanceCode = new_str[84:85].strip() + '\t'
                    scheduleCode = new_str[87:87].strip() + '\t'
                    pickUpCode = new_str[90:91].strip() + '\t'
                    periodFrom = new_str[94:100].strip() + '\t'
                    periodTo = new_str[101:107].strip() + '\t'
                    fuelType = new_str[109:110].strip() + '\t'
                    spcClm = new_str[114:115].strip() + '\t'
                    localUsage = new_str[118:118].strip() + '\t'
                    voucherPayment = new_str[179:180].strip() + '\t'
                    hhSize = new_str[186:188].strip() + '\t'
                    caseIncome = new_str[189:199].strip() + '\t'
                    heapVendorId = new_str[250:257].strip() + '\t'
                    customerAcctNo = new_str[270:285].strip() + '\t'
                    heapNominalBenefit = ' ' + '\t'
                    caseType = new_str[309:311].strip()

                    out_str = (caseNumber + caseName + caseAddress + caseCity + caseState + caseZip + localOffice + caseUnit + caseWorker + 
                        localActionCode + payType + caseMethod + caseAmount + issuanceCode + scheduleCode + pickUpCode + periodFrom + periodTo + fuelType +
                        spcClm + localUsage + voucherPayment + hhSize + caseIncome + heapVendorId + customerAcctNo + heapNominalBenefit + caseType)

            # write to results file                                
                    if out_str != ' ':
                        f = open(outfile,"a")
                        f.write(out_str + "\n")
                        f.close()

                        countWrites = countWrites +1 
                        countLines = 0
                        heapPerson = ' '



    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None,low_memory=False)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSBF630", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsa5257(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "INDIVIDUAL NAME\t" + "CIN\t" +
        "RETURN DATE")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()
               
        if line.startswith('LOCAL OFFICE'):
            localOffice = line[15:18].strip() + '\t'
            unit = line[37:42].strip() + '\t'
            worker = line[61:66].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*') and not line.startswith('QI-1') 
            and not line.startswith('PERIOD COVERED BY') and not line.startswith('REFERENCE') and not line.startswith('DISTRICT')
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER') and not line.startswith(escape) 
            and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[21:57].strip() + '\t'
                individualName = line[59:97].strip() + '\t'
                cin = line[99:107].strip() + '\t'
                returnDate = line[117:125].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + individualName + cin + returnDate)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSA5257", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc1025(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "PAY AMOUNT\t" + "PAY PERIOD")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[15:18].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('MASS AUTHORIZATION') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[23:55].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                caseType = line[56:58].strip() + '\t'
                payAmount = line[67:75].strip() + '\t'
                payPeriod = line[82:101].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + payee + caseType + payAmount + payPeriod)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC1025/1040", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc1026(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    if f_name.startswith('001-X-14-WMSC1026')  :
        outfilexlsx = "WINR1026 - MASS AUTH OF SUPPLEMENTAL SNAP - EXCEPTION [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
        out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "EXCEPTION REASON")
    if f_name.startswith('001-X-14-WMSC1041') :
        outfilexlsx = "WINR1041 - MASS AUTH OF SNAP - EXCEPTION REPORT [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
        out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "EXCEPTION REASON")
    if f_name.startswith('001-X-14-WMSC1661') :
        outfilexlsx = "WINR1661 - AUTH OF HEAP J9 SUPPL. BENEFITS - EXCEPTION [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
        out_str = ("Local Office\t" + "Unit\t" + "Worker\t" + "Case Number\t" + "Case Name\t" + "Case Type\t" +
        "Exception Reason")

    #outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

#    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
#        "EXCEPTION REASON")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[14:18].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('MASS AUTHORIZATION') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[13:45].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                caseType = line[46:48].strip() + '\t'
                exceptionReason = line[68:].strip()

                if f_name.startswith('001-X-14-WMSC1661') :
                    out_str = (localOffice + unit + worker + caseNumber + caseName + caseType + exceptionReason)
                else :
                    out_str = (localOffice + unit + worker + caseNumber + caseName + payee + caseType + exceptionReason)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC1026/1041", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc2027(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WINR2027 MASS ISSUANCE OF P-EBT - ELIGIBLE REPORT [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "CASE TYPE\t" +
        "PAY AMOUNT\t" + "PAY PERIOD FROM\t" + "PAY PERIOD TO\t" + "AVAIL DT")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[15:18].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('MASS ISSUANCE') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[23:55].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                caseType = line[56:58].strip() + '\t'
                payAmount = line[67:75].strip() + '\t'
                payPeriodFrom = line[82:90].strip() + '\t'
                payPeriodTo = line[93:101].strip() + '\t'
                availDt = line[107:117].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + payee + caseType + payAmount + payPeriodFrom
                    + payPeriodTo + availDt)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC2027", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsbhigh(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WMSBHIGH - RFI HIGH RISK LISTING [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE TYPE\t" + "CASE NUMBER\t" + "CASE NAME\t" + "Payee\t" + "INCOME\t" +
        "RESOURCE")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('REPORT') and not line.startswith('*')
            and not line.startswith('DISTRICT') and not line.startswith('OFFICE') 
            and not line.startswith('RFI HIGH') and not line.startswith('UNRESOLVED')
            and not line.startswith('SORT SEQUENCE') and not line.startswith('PLEASE') 
            and not line.startswith('THE ATTACHED') and not line.startswith('THE UNRESOLVED')
            and not line.startswith('FOR QUESTIONS') and not line.startswith('ASSISTANCE')
            and not line.startswith('HITS  THAT') and not line.startswith('REVENUE')
            and not line.startswith('BUSINESS') and not line.startswith('45 DAYS')
            and not line.startswith('"HITS"') and not line.startswith('OF THE WRM.')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            if countLines == 1  :
                localOffice = record[5:8].strip() + '\t'
                unit = record[14:19].strip() + '\t'
                worker = record[24:29].strip() + '\t'
                caseType = record[36:38].strip() + '\t'
                caseNumber = record[52:62].strip() + '\t'
                caseName = record[66:96].strip() + '\t'
                if '/PAYEE' in caseName :
                    caseName = caseName.replace('/PAYEE','')
                    payee = 'TRUE\t'
                else :
                    payee = '\t'
                income = record[98:105].strip() + '\t'
                resource = record[110:120].strip()

                out_str = (localOffice + unit + worker + caseType + caseNumber + caseName + payee + income + resource)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC2027", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc4210(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WINR4210 HEAP AUTO-CLOSE EXCEPTION REPORT [" + os.path.splitext(f_name[28:38])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "EXCEPTION REASON")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('REPORT DATE') and not line.startswith('*')
            and not line.startswith('DISTRICT') and not line.startswith('CASE NAME') 
            and not line.startswith('TOTAL CASES') and not line.startswith('END OF REPORT')
            and not line.startswith('REFERENCE NO')  
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            if countLines == 1  :
                localOffice = record[14:17].strip() + '\t'
                unit = record[38:43].strip() + '\t'
                worker = record[58:63].strip() + '\t'

            if countLines == 2  :
                caseName = record[0:32].strip() + '\t'
                caseNumber = record[32:42].strip() + '\t'
                exceptionReason = record[50:].strip()
#                if '/PAYEE' in caseName :
#                    caseName = caseName.replace('/PAYEE','')
#                    payee = 'TRUE\t'
#                else :
#                    payee = '\t'

                out_str = (localOffice + unit + worker + caseName + caseNumber + exceptionReason)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WINR4210", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def winr0799(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

#    outfilexlsx = "ec_" + os.path.splitext(f_name)[0] + ".xlsx"
    outfilexlsx = "WINR0799 Upstate SSN Auto Close Monthly Exception.xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("UNIT\t" + "LOCAL OFFICE\t" + "WORKER\t" + "CASE NUMBER\t" + "CASE NAME\t" + "CASE TYPE\t" + "CID ID\t" + "VR IND\t" +
        "Date of Death\t" + "Exception Message (Case is not closed)")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        #line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('NEW YORK STATE') and not line.startswith('Upstate Welfare')
            and not line.startswith('UPSTATE AUTO') and not line.startswith('Exception List') 
            and not line.startswith('REPORT DATE') and not line.startswith('WRTS REPORT')
            and not line.startswith('THIS REPORT') and not line.startswith('CONFIDENTIAL')
            and not line.startswith('FOR RESTRICTED') and not line.startswith('USE ONLY')
            and not line.startswith('District') and not line.startswith('Case Number')
            and not line.startswith('(Case is not') and not line.startswith('Total') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            #if countLines == 1  :
            if line.startswith('Unit')  :
                localOffice = record[22:25].strip() + '\t'
                unit = record[26:31].strip() + '\t'
                worker = record[32:37].strip() + '\t'

            #if countLines == 2  :
            else  :
                if record[8:57].strip() == ''  :
                    vrInd = record[82:84].strip() + '\t'
                    exceptionMessage = record[104:].strip()
                else  :
                    caseNumber = record[8:18].strip() + '\t'
                    caseName = record[27:57].strip() + '\t'
                    caseType = record[57:59].strip() + '\t'
                    cinId = record[67:75].strip() + '\t'
                    vrInd = record[82:84].strip() + '\t'
                    dateOfDeathMM = record[90:92].strip() + '/'
                    dateOfDeathDD = record[92:94].strip() + '/'
                    dateOfDeathYYYY = record[94:98].strip()
                    if dateOfDeathYYYY == ''  :
                        dateOfDeath = '\t'
                    else  :
                        dateOfDeath = dateOfDeathMM + dateOfDeathDD + dateOfDeathYYYY + '\t'
                    exceptionMessage = record[104:].strip()

                out_str = (unit + localOffice + worker + caseNumber + caseName + caseType + cinId + 
                    vrInd + dateOfDeath + exceptionMessage)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WINR0799", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wbb101(f_name):

    escape = '\u001B'
    carriage = '\r'
    colon = '\ua789'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "FBI FS AND PA" + colon + " FS EXTRACT" + colon + " FEDERAL BENEFIT INCREASE [" + os.path.splitext(f_name[30:40])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    out_str = ("LOCAL OFFICE\t" + "UNIT\t" + "WORKER\t" + "CASE NAME\t" + "CASE NUMBER\t" + "Mailing Address Line 1\t" +
        "Mailing Address City\t" + "Mailing Address State\t" + "Mailing Address Zip\t" + "SI Benefits\t" + "Veterans Benefits\t" +
        "Railroad Retirement Benefits\t" + "SS Benefits\t" + "EST EXEC")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        #line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('1234567890') and not line.startswith('File Sequence')
            and not line.startswith('WW')
            and not line.startswith('Product') and not line.startswith('Printer') 
            and not line.startswith('-Host') and not line.startswith('ALBEU')
            and not line.startswith('Host Full') and not line.startswith('Data File')
            and not line.startswith('Characters') and not line.startswith('Lines Per')
            and not line.startswith('Top Margin') and not line.startswith('Bottom Margin')
            and not line.startswith('Warning:') and not line.startswith('legally') 
            and not line.startswith('groups') and not line.startswith('government')
            and not line.startswith('constitutes') and not line.startswith('notice') 
            and not line.startswith('REPORT') and not line.startswith('WELFARE')
            and not line.startswith('FS EXTRACT') and not line.startswith('DISTRICT')
            and not line.startswith('CASE NAME')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            #if countLines == 1  :
            if line.startswith('OFFICE')  :
                localOffice = record[13:18].strip() + '\t'
                unit = record[24:29].strip() + '\t'
                worker = record[39:44].strip() + '\t'

            #if countLines == 2  :
            else  :
                if record[6:36].strip() != ''  :
                    if record[5:7].strip() == '' :
                        caseName = record[7:37].strip() + '\t'
                        caseNumber = record[41:53].strip() + '\t'
                        mailingAddressLine1 = record[56:96].strip() + '\t'
                        siBenefits = record[98:100].strip() + '\t'
                        veteranBenefits = record[101:103].strip() + '\t'
                        railroadRetirementBenefits = record[104:106].strip() + '\t'
                        ssBenefits = record[107:110].strip() + '\t'
                        estExec = record[112:].strip()
                    else :    
                        caseName = record[6:36].strip() + '\t'
                        caseNumber = record[40:52].strip() + '\t'
                        mailingAddressLine1 = record[55:95].strip() + '\t'
                        siBenefits = record[97:99].strip() + '\t'
                        veteranBenefits = record[100:102].strip() + '\t'
                        railroadRetirementBenefits = record[103:105].strip() + '\t'
                        ssBenefits = record[106:109].strip() + '\t'
                        estExec = record[111:].strip()
                else  :
                    mailingAddressCity = record[56:72].strip() + '\t'
                    mailingAddressState = record[73:75].strip() + '\t'
                    mailingAddressZip = record[77:82].strip() + '\t'

                    out_str = (localOffice + unit + worker + caseName + caseNumber + mailingAddressLine1 + mailingAddressCity + 
                        mailingAddressState + mailingAddressZip + siBenefits + veteranBenefits + railroadRetirementBenefits +
                        ssBenefits + estExec)

    # write to results file                                
                    if out_str != ' ':
                        f = open(outfile,"a")
                        f.write(out_str + "\n")
                        f.close()
                        countWrites = countWrites +1 
                        countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WBB101", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def lfrr0011(f_name):

    import openpyxl as pxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    escape = '\u001B'
    carriage = '\r'

    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    startDetail = True
    startSummary = False
    getFileName = False

    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)
    outfile2 = "ec2_" + f_name
    if os.path.exists(outfile2):
        os.remove(outfile2)

    file = open(f_name)
    lines = file.readlines()
    file.close()

    out_str = ("CHECK/SERV PERIOD Start\t" + "CHECK/SERV PERIOD End\t" + "CHECK NUMBER\t" + "Check Date\t" + "Payment Amount\t" + "Prior Year Amount\t" +
        "Current Year Amount\t" + "Category\t" + "Check Status\t" + "Cancel Date\t" + "Check Type\t" + "Case #\t" + "Adjustment Memo")
    out_str2 = ("CATEGORY\t" + "Prior Year ISSUE AMOUNT\t" + "Prior Year CANCEL AMOUNT\t" + "Prior Year BALANCE\t" + 
        "CURRENT YEAR ISSUE AMOUNT\t" + "CURRENT YEAR CANCEL AMOUNT\t" + "CURRENT YEAR BALANCE")    

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    if out_str2 != ' ':
        f = open(outfile2,"a")
        f.write(out_str2 + "\n")
        f.close()
    
    for line in lines:
        record = line
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('REPORT PERIOD') and getFileName == False :
            outfilexlsx = "LFRR0011 PAYMENTS FOR PRIOR YEAR SERVICES [" + record[15:34].replace('/','') + "].xlsx"
            getFileName = True
            if os.path.exists(outfilexlsx):
                os.remove(outfilexlsx)

        if line.startswith('CATEGORY') :
            startSummary = True

        if (not line.startswith('REPORT') and not line.startswith('BICS') and not line.startswith('SORTED')
            and not line.startswith('PAYMENTS') and not line.startswith('CHECK') and not line.startswith('*****')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startSummary == False):

            countLines = countLines + 1

#            if countLines == 1  :
            checkServPeriodStart = record[0:8].strip() + '\t'
            checkServPeriodEnd = record[11:19].strip() + '\t'
            checkNumber = record[21:33].strip() + '\t'
            checkDate = record[35:43].strip() + '\t'
            paymentAmount = record[44:52].strip() + '\t'
            priorYearAmount = record[53:61].strip() + '\t'
            currentYearAmount = record[62:70].strip() + '\t'
            category = record[72:81].strip() + '\t'
            checkStatus = record[82:85].strip() + '\t'
            cancelDate = record[87:95].strip() + '\t'
            checkType = record[97:100].strip() + '\t'
            caseNumber = record[102:113].strip() + '\t'
            adjustmentMemo = record[114:].strip()



            out_str = (checkServPeriodStart + checkServPeriodEnd + checkNumber + checkDate + paymentAmount + priorYearAmount +
                currentYearAmount + category + checkStatus + cancelDate + checkType + caseNumber + adjustmentMemo)

            # write to results file                                
            if out_str != ' ':
                f = open(outfile,"a")
                f.write(out_str + "\n")
                f.close()
                countWrites = countWrites +1 
                countLines = 0

        if (not line.startswith('TOTALS') and not line.startswith('CATEGORY')
            and not line.startswith(escape) and not line.startswith(carriage) 
            and len(line) != 0 and startSummary == True):

            countLines = countLines + 1

#            if countLines == 1  :
            sumCategory = record[0:8].strip() + '\t'
            sumPriorYearIssueAmount = record[12:24].strip() + '\t'
            sumPriorYearCancelAmount = record[28:40].strip() + '\t'
            sumPriorYearBalance = record[46:58].strip() + '\t'
            sumCurrentYearIssueAmount = record[83:95].strip() + '\t'
            sumCurrentYearCancelAmount = record[99:111].strip() + '\t'
            sumCurrentYearBalance = record[116:128].strip()

            out_str2 = (sumCategory + sumPriorYearIssueAmount + sumPriorYearCancelAmount + sumPriorYearBalance + 
                sumCurrentYearIssueAmount + sumCurrentYearCancelAmount + sumCurrentYearBalance)

            # write to results file                                
            if out_str2 != ' ':
                f = open(outfile2,"a")
                f.write(out_str2 + "\n")
                f.close()
                countWrites = countWrites +1 
                countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Detail',index=False,header=False)               
    os.remove(outfile)            

    excel_book = pxl.load_workbook(outfilexlsx)

    excel_book.create_sheet('Summary')
    
    df2 = pd.read_csv(outfile2,sep='\t',lineterminator='\n',header=None)
    rows = dataframe_to_rows(df2,index=False,header=False)
    ws = excel_book['Summary']
    for r_idx, row in enumerate(rows, 1) :
        for c_idx, value in enumerate(row,1) :
            ws.cell(row=r_idx, column=c_idx,value=value)
    os.remove(outfile2)            
    excel_book.save(outfilexlsx)
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - LFRR0011", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def retrt010(f_name):

    escape = '\u001B'
    carriage = '\r'
    colon = '\ua789'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "RETRT010 - RETRO RATE CHANGE APPROVED VOUCHERS REPORT [" + os.path.splitext(f_name[25:35])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    out_str = ("Voucher #\t" + "Vendor\t" + "Vendor Name\t" + "Placement Vendor ID\t" + "Remitance (Payment Amount of new Retro Rate Voucher line)\t" + 
        "Line #\t" + "Case #\t" + "Recipient\t" + "Dob\t" + "New LOD (Level of Difficulty)\t" + "New Program Type (PG)\t" +
        "New Service Type\t" + "New Service Period From\t" + "New Service Period To\t" + "New POS Type Pass-Through Rate (PASS)\t" +
        "New Clothing Rate\t" + "New Administrative Rate\t" + "New Total Rate\t" + "Original Voucher\t" +
        "Check #\t" + "Check Date\t" + "Auth #\t" + "Original LOD (Level of Difficulty)\t" + "Original Program Type (PG)\t" +
        "Original Service Type\t" + "Original Service Period From\t" + "Original Service Period To\t" + "Original POS Type Pass-Through Rate (PASS)\t" + 
        "Original Clothing Rate\t" + "Original Administrative Rate\t" + "Original Total Rate")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()
    
    for line in lines:
        record = line
        #line = line.strip().upper()
        line = line.strip()

        if (not line.startswith('REPORT DATE') and not line[40:45] == ('RETRO')
            and not line.startswith('BICS') and not line.startswith('VOUCHER AMOUNT')
            and not line.startswith('LINE')
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0):

            countLines = countLines + 1

            #if countLines == 1  :
            if line.startswith('VOUCHER NO')  :
                voucherNumber = record[14:22].strip() + '\t'
                vendor = record[32:40].strip() + '\t'
                vendorName = record[55:84].strip() + '\t'
                placementVendorId = record[103:111].strip() + '\t'

            #if countLines == 2  :
            else  :
                if record[6:36].strip() != ''  :
                    if record[0:1].strip() == '' :
                        lineNumber = record[1:5].strip() + '\t'
                        caseNumber = record[6:16].strip() + '\t'
                        recipientId = record[17:25].strip() + '\t'
                        dob = record[26:34].strip() + '\t'
                        newLd = record[35:37].strip() + '\t'
                        newPg = record[38:40].strip() + '\t'
                        newSt = record[41:43].strip() + '\t'
                        newServicePeriodFrom = record[44:52].strip() + '\t'
                        newServicePeriodTo = record[53:61].strip() + '\t'
                        newPass = record[93:100].strip() + '\t'
                        newCloth = record[101:108].strip() + '\t'
                        newAdmin = record[109:116].strip() + '\t'
                        newRate = record[117:124].strip() + '\t'
                        newRemit = record[125:132].strip() + '\t'
                    else :    
                        origVoucher = record[17:25].strip() + '\t'
                        origLd = record[35:37].strip() + '\t'
                        origPg = record[38:40].strip() + '\t'
                        origSt = record[41:43].strip() + '\t'
                        origServicePeriodFrom = record[44:52].strip() + '\t'
                        origServicePeriodTo = record[53:61].strip() + '\t'
                        
                        checkNumber = record[62:69].strip() + '\t'
                        checkDate = record[75:83].strip() + '\t'
                        authNumber = record[84:92].strip() + '\t'
                        
                        origPass = record[93:100].strip() + '\t'
                        origCloth = record[101:108].strip() + '\t'
                        origAdmin = record[109:116].strip() + '\t'
                        origRate = record[117:124].strip()

                        out_str = (voucherNumber + vendor + vendorName + placementVendorId + newRemit + lineNumber +
                            caseNumber + recipientId + dob + newLd + newPg + newSt + newServicePeriodFrom + newServicePeriodTo + 
                            newPass + newCloth + newAdmin + newRate + origVoucher + checkNumber + checkDate + authNumber +
                            origLd + origPg + origSt + origServicePeriodFrom + origServicePeriodTo + origPass + origCloth +
                            origAdmin + origRate)

    # write to results file                                
                        if out_str != ' ':
                            f = open(outfile,"a")
                            f.write(out_str + "\n")
                            f.close()
                            countWrites = countWrites +1 
                            countLines = 0

    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - RETRT010", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

def wmsc1047(f_name):

    escape = '\u001B'
    carriage = '\r'
            
    # setup count variables
    countLines = 0
    countWrites = 0
    countFiles = 0
    
    # read file
    outfile = "ec_" + f_name
    if os.path.exists(outfile):
        os.remove(outfile)

    outfilexlsx = "WINR1047 AUTH OF REPLACEMENT SNAP - EXCEPTION [" + os.path.splitext(f_name[28:39])[0] + "].xlsx"
    if os.path.exists(outfilexlsx):
        os.remove(outfilexlsx)

    file = open(f_name)
    lines = file.readlines()
    file.close()
            
    getVendorName = False
    startDetail = False

    out_str = ("Office\t" + "Unit\t" + "Worker\t" + "Case Number\t" + "Case Name\t" + "Case Type\t" + "Exception Reason")

    print("f_name is " + f_name)

    # write to results file                                
    if out_str != ' ':
        f = open(outfile,"a")
        f.write(out_str + "\n")
        f.close()

    # setup variables
    heapPerson = ' '
    
    for line in lines:
        line = line.strip().upper()
        line = line.strip()

        if line.startswith('DISTRICT'):
            unit = line[51:56].strip() + '\t'
            startDetail = False

        if line.startswith('LOCAL OFFICE'):
            localOffice = line[14:17].strip() + '\t'
            worker = line[53:58].strip() + '\t'
            startDetail = False

        if line.startswith('CASE NUMBER'):
                    startDetail = True

        if (not line.startswith('REPORT') and not line.startswith('*') 
            and not line.startswith('ERIE AUTHORIZATION') and not line.startswith('DISTRICT') 
            and not line.startswith('LOCAL OFFICE') and not line.startswith('CASE NUMBER')
            and not line.startswith('WMS REPORT') and not line.startswith('REFERENCE') 
            and not line.startswith('END OF REPORT') and not line.startswith('TOTAL') 
            and not line.startswith(escape) and not line.startswith(carriage) and len(line) != 0 and startDetail == True):

            countLines = countLines + 1

            if countLines == 1  :
                caseNumber = line[:10].strip() + '\t'
                caseName = line[13:45].strip() + '\t'
#                if '/PAYEE' in caseName :
#                    caseName = caseName.replace('/PAYEE','')
#                    payee = 'TRUE\t'
#                else :
#                    payee = '\t'
                caseType = line[46:48].strip() + '\t'
                exceptionReason = line[68:].strip()

                out_str = (localOffice + unit + worker + caseNumber + caseName + caseType + exceptionReason)

    # write to results file                                
                if out_str != ' ':
                    f = open(outfile,"a")
                    f.write(out_str + "\n")
                    f.close()
                    countWrites = countWrites +1 
                    countLines = 0
                    heapPerson = ' '
              
    df = pd.read_csv(outfile,sep='\t',lineterminator='\n',header=None)
    df.to_excel(outfilexlsx,'Sheet1',index=False,header=False)   
    os.remove(outfile)            
    print("Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    messagebox.showinfo("Completed - WMSC1025/1040", "Completed!  " + str(countWrites) + " records converted from " + f_name + " !")

    return

#===================================================================================================

main()

